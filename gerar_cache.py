"""
Gera o cache de resumo dos jogos (games-summary-wN.json) localmente,
replicando a MESMA lógica de cálculo do backend, sem depender do Flask.

Uso: python gerar_cache.py 4
"""
import sys
import re
import json
from pathlib import Path
from datetime import datetime
from difflib import SequenceMatcher
import openpyxl

# Fonte dos dados (mesma pasta usada localmente pelo backend)
dev_path = Path("/Users/luizprado/Downloads/Claude/Campeonato Petz")
BASE_PATH = dev_path if dev_path.exists() else Path(__file__).parent / "data"
SEMANA_ANTERIOR = BASE_PATH / "SEMANA ANTERIOR"
SEMANA_ATUAL = BASE_PATH / "SEMANA ATUAL"
SIMILARIDADE_MIN = 0.6
FILE_ALIASES = {}


def _listar_xlsx(semana_path):
    if not semana_path.exists():
        return []
    return sorted(f for f in semana_path.glob("*.xlsx") if not f.name.startswith("~"))


def _chave(nome_arquivo):
    base = nome_arquivo.rsplit(".", 1)[0].upper()
    return re.sub(r"[^A-Z0-9]", "", base)


def _similaridade(nome_a, nome_b):
    return SequenceMatcher(None, _chave(nome_a), _chave(nome_b)).ratio()


def mapear_indicadores():
    atual_files = _listar_xlsx(SEMANA_ATUAL)
    anterior_files = _listar_xlsx(SEMANA_ANTERIOR)
    indicadores = {}
    for af in atual_files:
        indicadores[af.name] = {"anterior": None, "atual": af}
    restantes = list(anterior_files)
    for canonico, slots in indicadores.items():
        for pf in list(restantes):
            if pf.name == canonico or FILE_ALIASES.get(pf.name) == canonico:
                slots["anterior"] = pf
                restantes.remove(pf)
                break
    for canonico, slots in indicadores.items():
        if slots["anterior"] is not None:
            continue
        melhor, melhor_score = None, 0.0
        for pf in restantes:
            score = _similaridade(canonico, pf.name)
            if score > melhor_score:
                melhor, melhor_score = pf, score
        if melhor and melhor_score >= SIMILARIDADE_MIN:
            slots["anterior"] = melhor
            restantes.remove(melhor)
    for pf in restantes:
        indicadores.setdefault(pf.name, {"anterior": pf, "atual": None})
    return indicadores


def ler_dias_loja(file_path, sigla):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    header = list(ws.iter_rows(max_row=1, values_only=True))[0]
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            continue
        if row[0] == sigla:
            dias = {}
            for col_idx in range(2, len(header)):
                if header[col_idx] and "202" in str(header[col_idx]):
                    dia_nome = str(header[col_idx]).split("(")[1].rstrip(")")
                    valor = row[col_idx]
                    try:
                        if valor == "-" or valor is None:
                            dias[dia_nome] = 0
                        else:
                            dias[dia_nome] = round(float(valor), 2)
                    except (ValueError, TypeError):
                        dias[dia_nome] = 0
            return dias
    return None


def calcularPlacarBackend(team1, team2, semana, hojeIdx=None, _mapa=None):
    try:
        mapa = _mapa if _mapa is not None else mapear_indicadores()
        score1 = 0
        score2 = 0
        diasOrdenados = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']
        for arquivo, slots in mapa.items():
            dados1_ind = {}
            dados2_ind = {}
            for semana_type in ("anterior", "atual"):
                file_path = slots.get(semana_type)
                if not file_path:
                    continue
                dias1 = ler_dias_loja(file_path, team1)
                dias2 = ler_dias_loja(file_path, team2)
                if dias1:
                    dados1_ind[semana_type] = dias1
                if dias2:
                    dados2_ind[semana_type] = dias2
            if not dados1_ind or not dados2_ind:
                continue
            dias1Anterior = dados1_ind.get('anterior', {})
            dias1Atual = dados1_ind.get('atual', {})
            dias2Anterior = dados2_ind.get('anterior', {})
            dias2Atual = dados2_ind.get('atual', {})
            diasAcontar = diasOrdenados[:hojeIdx+1] if hojeIdx is not None else diasOrdenados
            total1Anterior = sum(dias1Anterior.get(dia, 0) for dia in diasAcontar)
            total1Atual = sum(dias1Atual.get(dia, 0) for dia in diasAcontar)
            total2Anterior = sum(dias2Anterior.get(dia, 0) for dia in diasAcontar)
            total2Atual = sum(dias2Atual.get(dia, 0) for dia in diasAcontar)
            evolucao1 = (total1Atual - total1Anterior)
            evolucao2 = (total2Atual - total2Anterior)
            if evolucao1 > evolucao2:
                score1 += 1
            elif evolucao2 > evolucao1:
                score2 += 1
        return score1, score2
    except Exception as e:
        print(f"Erro ao calcular placar {team1} vs {team2}: {e}")
        return 0, 0


def gerar(semana):
    confrontos_path = BASE_PATH / "Confrontos" / f"Semana {semana}.xlsx"
    if not confrontos_path.exists():
        print(f"❌ Confrontos não encontrados: {confrontos_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(confrontos_path, data_only=True)
    ws = wb.active
    confrontos = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx <= 1:
            continue
        team1 = row[2]
        team2 = row[4]
        if team1 and team2:
            confrontos.append({"team1": team1, "team2": team2})

    print(f"⏳ Calculando {len(confrontos)} jogos da semana {semana}...")
    hoje_idx = (datetime.now().weekday() + 1) % 7
    mapa = mapear_indicadores()  # calcula uma vez, reutiliza

    games_summary = []
    for idx, conf in enumerate(confrontos):
        team1 = conf["team1"]
        team2 = conf["team2"]
        score1_proj, score2_proj = calcularPlacarBackend(team1, team2, semana, None, mapa)
        score1_acum, score2_acum = calcularPlacarBackend(team1, team2, semana, hoje_idx, mapa)
        games_summary.append({
            "team1": team1,
            "team2": team2,
            "scoreProjected": f"{score1_proj} x {score2_proj}",
            "scoreAccumulated": f"{score1_acum} x {score2_acum}",
            "hojeIdx": hoje_idx,
        })
        if (idx + 1) % 20 == 0:
            print(f"  {idx + 1}/{len(confrontos)} calculados...")

    data = {
        "week": semana,
        "lastUpdated": datetime.now().isoformat(),
        "total": len(games_summary),
        "games": games_summary,
    }
    cache_dir = Path(__file__).parent / "data" / "cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    out = cache_dir / f"games-summary-w{semana}.json"
    with open(out, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"✅ {len(games_summary)} jogos salvos em {out}")


if __name__ == "__main__":
    semana = int(sys.argv[1]) if len(sys.argv) > 1 else 4
    gerar(semana)
