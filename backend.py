"""
Backend Flask - Campeonato Petz 2026
Integração com dados de Semana Anterior e Semana Atual
"""

from flask import Flask, jsonify, request, send_from_directory, session
from flask_cors import CORS
from flask_login import login_user, logout_user, login_required, current_user
import shutil
import openpyxl
import os
import re
import requests
from datetime import datetime, timedelta, timezone
from difflib import SequenceMatcher
from pathlib import Path
from auth import (db, login_manager, Usuario, Acesso, init_db,
                  autenticar_emergencia, invalidar_cache_usuarios)

app = Flask(__name__)
CORS(app)

# ------------------------------------------------------------------
# Banco de dados de usuários
# ------------------------------------------------------------------
# Prioriza um Postgres gerenciado (Vercel/Neon), que é PERMANENTE. Sem ele,
# cai para SQLite — que em /tmp no Vercel é efêmero (usuários/senhas se perdem
# a cada reinício da instância).
def _database_uri():
    for var in ('POSTGRES_URL_NON_POOLING', 'POSTGRES_URL', 'DATABASE_URL',
                'POSTGRES_PRISMA_URL'):
        url = os.environ.get(var)
        if url:
            # SQLAlchemy exige o driver explícito
            if url.startswith('postgres://'):
                url = url.replace('postgres://', 'postgresql://', 1)
            if url.startswith('postgresql://'):
                url = url.replace('postgresql://', 'postgresql+psycopg://', 1)
            # o parâmetro do Prisma não é aceito pelo driver
            url = url.replace('?pgbouncer=true', '?').replace('&pgbouncer=true', '')
            print(f"🗄️  Usando Postgres permanente (via {var})")
            return url
    if os.environ.get('VERCEL') or not os.access(str(Path(__file__).parent), os.W_OK):
        print("⚠️  Sem Postgres configurado — usando SQLite em /tmp (efêmero!)")
        return 'sqlite:////tmp/campeonato.db'
    return f"sqlite:///{Path(__file__).parent / 'campeonato.db'}"


app.config['SQLALCHEMY_DATABASE_URI'] = _database_uri()
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'pool_pre_ping': True}
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'sua-chave-secreta-mude-isso-em-producao')
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = False  # Mude para True em produção (HTTPS)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)

db.init_app(app)
login_manager.init_app(app)
login_manager.login_view = 'login'

# Inicializar banco de dados
init_db(app)


@app.after_request
def _sem_cache_api(resp):
    """Impede navegador/CDN de servir respostas de API em cache (dados mudam)."""
    if request.path.startswith('/api/'):
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
    return resp

# Configuração de caminhos
# Em desenvolvimento local: /Users/luizprado/Downloads/Claude/Campeonato Petz
# Em Vercel: use caminhos relativos
try:
    dev_path = Path("/Users/luizprado/Downloads/Claude/Campeonato Petz")
    if dev_path.exists():
        BASE_PATH = dev_path
    else:
        BASE_PATH = Path(__file__).parent / "data"
except:
    BASE_PATH = Path(__file__).parent / "data"

# Base "empacotada" (repositório/dev). Confrontos vêm sempre daqui.
BUNDLED_BASE = BASE_PATH

# Após um reprocessamento, os dados frescos do SharePoint são gravados em /tmp
# (único diretório gravável no Vercel). As leituras passam a preferir /tmp.
TMP_BASE = Path('/tmp/campeonato_data')


def active_base():
    """Retorna a base de dados ativa: /tmp se já houve download do SharePoint,
    senão a base empacotada no repositório.

    Considera TAMBÉM as subpastas por rodada — com a estrutura nova os arquivos
    ficam em 'SEMANA ATUAL/rodada N' e a raiz fica vazia; sem isso o app cairia
    para a cópia antiga do repositório.
    """
    for pasta in ("SEMANA ATUAL", "SEMANA ANTERIOR"):
        p = TMP_BASE / pasta
        if p.exists() and (any(p.glob("*.xlsx")) or any(p.glob("rodada */*.xlsx"))):
            return TMP_BASE
    return BUNDLED_BASE


def _tem_xlsx(p):
    return p.exists() and any(p.glob("*.xlsx"))


def pasta_dados(nome):
    """Pasta de apoio (Estrutura, Historico, ...) preferindo o que veio do
    SharePoint em /tmp e caindo para a cópia do repositório.

    Sem esse fallback, um throttle do SharePoint zerava estrutura, histórico e
    classificação — mesmo com os arquivos empacotados no deploy.
    """
    tmp = TMP_BASE / nome
    if _tem_xlsx(tmp):
        return tmp
    return BUNDLED_BASE / nome


def _rodada_iniciada(base, n):
    """A rodada n já tem base para ser exibida?

    Basta a SEMANA ANTERIOR: no primeiro dia da rodada só existe a base de
    comparação, e é assim que deve aparecer — semana anterior com os valores
    e semana atual zerada, evoluindo conforme os dias entram.

    Olha as DUAS cópias (SharePoint e empacotada). Se olhasse só a ativa, uma
    rodada que chegou pela cópia empacotada — porque a sincronização do
    SharePoint travou — seria considerada "não iniciada" e o app cairia para a
    rodada anterior, mostrando os indicadores da semana passada.
    """
    for b in {base, TMP_BASE, BUNDLED_BASE}:
        if (_tem_xlsx(b / "SEMANA ATUAL" / f"rodada {n}")
                or _tem_xlsx(b / "SEMANA ANTERIOR" / f"rodada {n}")):
            return True
    return False


def rodada_efetiva(semana=None):
    """Qual rodada realmente tem dados para exibir.

    Ordem: a própria rodada -> estrutura antiga (arquivos na raiz) -> a rodada
    anterior mais recente que tenha dados. Retorna (rodada, usou_raiz).
    Isso mantém o app útil quando a rodada já começou (confrontos publicados)
    mas as planilhas de venda dela ainda não subiram.
    """
    base = active_base()
    if semana is None:
        semana = semana_atual()
    if _rodada_iniciada(base, semana):
        return semana, False
    if _tem_xlsx(base / "SEMANA ATUAL"):
        return semana, True          # estrutura antiga (sem subpastas)
    for n in range(semana - 1, 0, -1):
        if _rodada_iniciada(base, n):
            return n, False
    return semana, True


# Pasta onde montamos a combinação do que veio do SharePoint com o que está
# empacotado no deploy. Fica em /tmp por ser o único lugar gravável no Vercel.
MERGE_BASE = Path('/tmp/campeonato_merge')
_merge_cache = {}


def _ultimo_dia_do_arquivo(caminho):
    """Índice do último dia com valor no arquivo (-1 se não tem nenhum).

    É a medida de "mais recente" que importa aqui: entre duas cópias do mesmo
    indicador, a boa é a que tem mais dias lançados. Data de modificação não
    serve — a cópia empacotada nasce com a data do deploy e ganharia sempre.
    """
    try:
        import calculo_rapido as cr
        dados = cr._carregar_arquivo(caminho)
        ultimo = -1
        for dias in dados.values():
            for i, d in enumerate(cr.DIAS_ORDENADOS):
                if dias.get(d):
                    ultimo = max(ultimo, i)
        return ultimo
    except Exception:
        return -1


def _combinar_semana(pasta, rod):
    """Junta as duas cópias de uma rodada, arquivo a arquivo, ficando com a
    mais adiantada de cada indicador.

    O SharePoint sincroniza um arquivo por vez e às vezes trava no meio: dá
    para ter VENDAS até domingo lá e MARCA ZEE até quarta, enquanto a cópia
    empacotada tem os dois até domingo. Escolher a pasta inteira não resolve
    (a do SharePoint "ganha" por causa do VENDAS); escolher arquivo a arquivo,
    sim. Empate fica com o SharePoint, que é a fonte viva.

    Devolve o diretório combinado, ou None se não valer a pena combinar.
    """
    origens = [b / pasta / f"rodada {rod}" for b in (TMP_BASE, BUNDLED_BASE)]
    tmp_dir, pkg_dir = origens
    com_arquivos = [d for d in origens if _tem_xlsx(d)]
    if not com_arquivos:
        return None
    if len(com_arquivos) == 1:
        # Só uma das cópias tem a rodada: usa ela direto, sem combinar.
        return com_arquivos[0]

    # Assinatura barata para não refazer a combinação a cada requisição.
    try:
        assinatura = tuple(sorted(
            (f.name, f.stat().st_mtime_ns, f.stat().st_size)
            for d in origens for f in d.glob("*.xlsx")))
    except OSError:
        return None
    destino = MERGE_BASE / pasta / f"rodada {rod}"
    if _merge_cache.get((pasta, rod)) == assinatura and destino.is_dir():
        return destino

    nomes = {f.name for d in origens for f in d.glob("*.xlsx")
             if not f.name.startswith("~")}
    if not nomes:
        return None
    escolhas = {}
    for nome in nomes:
        cands = [d / nome for d in origens if (d / nome).exists()]
        # max() com chave estável: em empate fica o primeiro, que é o /tmp.
        escolhas[nome] = max(cands, key=_ultimo_dia_do_arquivo)

    # Só vale combinar se alguma escolha vier da cópia empacotada.
    if all(c.is_relative_to(TMP_BASE) for c in escolhas.values()):
        _merge_cache[(pasta, rod)] = assinatura
        return None

    try:
        if destino.exists():
            shutil.rmtree(destino)
        destino.mkdir(parents=True, exist_ok=True)
        for nome, origem in escolhas.items():
            shutil.copy2(origem, destino / nome)
        vindos = sorted(n for n, c in escolhas.items() if not c.is_relative_to(TMP_BASE))
        print(f"🔀 {pasta}/rodada {rod}: usando a cópia empacotada para {', '.join(vindos)} "
              f"(mais adiantada que a do SharePoint).")
        _merge_cache[(pasta, rod)] = assinatura
        return destino
    except Exception as e:
        print(f"⚠️ Não consegui combinar {pasta}/rodada {rod}: {e}")
        return None


def _dir_semana(pasta, semana=None):
    """Diretório dos dados de uma semana.

    Estrutura preferida (por rodada):  <base>/SEMANA ATUAL/rodada 8
    Cai para a pasta antiga (<base>/SEMANA ATUAL) ou para a última rodada com
    dados, de modo que a migração dos arquivos possa ser feita aos poucos.

    Quando as duas cópias (SharePoint e empacotada) têm a rodada, entra a
    combinação arquivo a arquivo — ver _combinar_semana().
    """
    base = active_base()
    rod, usou_raiz = rodada_efetiva(semana)
    if usou_raiz:
        return base / pasta
    combinado = _combinar_semana(pasta, rod)
    return combinado or (base / pasta / f"rodada {rod}")


def dir_anterior(semana=None):
    return _dir_semana("SEMANA ANTERIOR", semana)


def dir_atual(semana=None):
    return _dir_semana("SEMANA ATUAL", semana)


def dir_confrontos():
    """Pasta de confrontos: prefere a baixada do SharePoint (/tmp), senão a
    empacotada no repositório.

    Para pegar UM arquivo específico use arquivo_confrontos(): a pasta escolhida
    aqui pode não ter a rodada mais nova.
    """
    tmp_conf = TMP_BASE / "Confrontos"
    if tmp_conf.exists() and any(tmp_conf.glob("Semana *.xlsx")):
        return tmp_conf
    return BUNDLED_BASE / "Confrontos"


def arquivo_confrontos(nome):
    """Caminho de um arquivo de confronto, olhando as duas cópias.

    semanas_disponiveis() já une o que existe no /tmp e no empacotado, então a
    rodada 10 pode aparecer no seletor vindo só da cópia empacotada. Sem olhar
    as duas aqui, o app anunciaria a rodada e depois não acharia o arquivo dela.
    Entre as duas cópias do mesmo arquivo, fica a maior — um confronto novo
    nunca tem menos jogos que o antigo.
    """
    cands = [b / "Confrontos" / nome for b in (TMP_BASE, BUNDLED_BASE)]
    existentes = [c for c in cands if c.exists()]
    if not existentes:
        return dir_confrontos() / nome
    return max(existentes, key=lambda c: c.stat().st_size)


def semanas_disponiveis():
    """Lista os números de semana com arquivo de confrontos disponível."""
    semanas = set()
    for base in (TMP_BASE / "Confrontos", BUNDLED_BASE / "Confrontos"):
        if not base.exists():
            continue
        for f in base.glob("Semana *.xlsx"):
            if f.name.startswith("~"):
                continue
            m = re.search(r"Semana\s+(\d+)", f.name)
            if m:
                semanas.add(int(m.group(1)))
    return sorted(semanas)


def semana_atual():
    """Semana vigente = maior número de semana com confrontos disponíveis."""
    s = semanas_disponiveis()
    return s[-1] if s else 4


# Mantidos por compatibilidade (usados só como base do Confrontos e afins)
SEMANA_ANTERIOR = BASE_PATH / "SEMANA ANTERIOR"
SEMANA_ATUAL = BASE_PATH / "SEMANA ATUAL"

# ------------------------------------------------------------------
# Frescor dos dados (resolve o /tmp por-instância do Vercel)
# ------------------------------------------------------------------
# Cada instância serverless tem seu próprio /tmp. Sem um controle de validade,
# uma instância que baixou os dados na quarta continua servindo quarta mesmo
# depois de alguém atualizar o SharePoint. Este TTL faz cada instância
# rebaixar os arquivos do SharePoint quando estão "velhos", convergindo todas
# para o dado mais recente em poucos minutos.
import time
import threading

# As planilhas mudam no máximo algumas vezes por dia; 90s de TTL fazia cada
# visita rebaixar ~90 arquivos do SharePoint e levou a 429 (throttle).
CACHE_TTL_SEGUNDOS = 900         # frescor máximo antes de rebaixar (15 min)
ESPERA_APOS_THROTTLE = 900       # depois de um 429, aguarda antes de tentar
RETENTAR_APOS_FALHA = 60         # download parcial: tenta de novo em 1 min
TTL_INCOMPLETO = 120             # indicadores em dias diferentes: rebaixa em 2 min
_fetch_lock = threading.Lock()
_MARKER = TMP_BASE / ".fetched_at"


def garantir_rodada(semana):
    """Garante que os dados da rodada pedida estejam em /tmp.
    Permite reabrir rodadas passadas: se a subpasta 'rodada N' ainda não foi
    baixada nesta instância, busca sob demanda (uma vez)."""
    if not semana:
        return
    alvo = TMP_BASE / "SEMANA ATUAL" / f"rodada {semana}"
    if alvo.exists() and any(alvo.glob("*.xlsx")):
        return
    marcador = TMP_BASE / f".rodada_{semana}_tentada"
    if marcador.exists():
        return   # já tentamos; a rodada provavelmente não usa subpastas
    with _fetch_lock:
        if alvo.exists() and any(alvo.glob("*.xlsx")):
            return
        try:
            import sharepoint
            print(f"⏳ Baixando dados da rodada {semana}...")
            sharepoint.baixar_rodada(semana, str(TMP_BASE), timeout=25)
        except Exception as e:
            print(f"⚠️ garantir_rodada({semana}) falhou: {e}")
        finally:
            try:
                TMP_BASE.mkdir(parents=True, exist_ok=True)
                marcador.write_text(str(time.time()))
            except Exception:
                pass


def _idade_tmp():
    """Idade (segundos) dos dados em /tmp; None se nunca baixado."""
    try:
        if _MARKER.exists():
            return time.time() - float(_MARKER.read_text().strip())
    except Exception:
        pass
    return None


def _em_espera_throttle():
    """Estamos no período de espera após um 429 do SharePoint?"""
    try:
        import sharepoint
        if not sharepoint.ULTIMO_THROTTLE:
            return False
        return (time.time() - sharepoint.ULTIMO_THROTTLE) < ESPERA_APOS_THROTTLE
    except Exception:
        return False


def _ttl_efetivo():
    """Quanto tempo o /tmp desta instância vale antes de baixar de novo.

    Normalmente CACHE_TTL_SEGUNDOS. Mas quando os indicadores da rodada estão
    em dias diferentes, o TTL cai para TTL_INCOMPLETO. O motivo é o jeito como
    o upload acontece: os arquivos são salvos um a um e o SharePoint sincroniza
    cada um no seu tempo, então uma instância que baixa no meio do caminho pega
    metade novo e metade velho — e ficaria congelada assim por 15 minutos,
    mostrando indicadores "pendentes" que já estão lá.

    Rebaixar mais cedo resolve esse caso. Quando o indicador está realmente
    atrasado (o usuário ainda não subiu), o custo é só uma consulta a mais de
    vez em quando.
    """
    try:
        import calculo_rapido as cr
        sem = semana_atual()
        memoria = cr.carregar_tudo(dir_anterior(sem), dir_atual(sem))
        ultimos = set()
        for _arq, s in memoria.items():
            dias = [d for d in cr.DIAS_ORDENADOS
                    if any((v or {}).get(d) for v in (s.get("atual") or {}).values())]
            ultimos.add(dias[-1] if dias else None)
        if len(ultimos) > 1:
            return TTL_INCOMPLETO
    except Exception:
        pass
    return CACHE_TTL_SEGUNDOS


def garantir_arquivos_frescos(force=False):
    """Garante que /tmp tenha os arquivos do SharePoint razoavelmente frescos.
    Baixa se nunca baixou, se passou do TTL, ou se force=True. Protegido por
    lock para evitar downloads simultâneos na mesma instância."""
    idade = _idade_tmp()
    ttl = _ttl_efetivo()
    if not force and idade is not None and idade < ttl:
        return  # já está fresco o suficiente
    if _em_espera_throttle() and active_base() == TMP_BASE:
        return  # throttle recente e já temos dados baixados: não insistir
    with _fetch_lock:
        # Re-checar após o lock: outra thread pode ter acabado de baixar
        idade = _idade_tmp()
        if not force and idade is not None and idade < ttl:
            return
        try:
            import sharepoint
            print("⏳ Atualizando dados do SharePoint (frescor)...")
            # Em duas fases: a rodada vigente é definida pelos Confrontos, que
            # só ficam corretos DEPOIS do primeiro download. Perguntar antes
            # devolveria a semana da cópia empacotada (desatualizada) e as
            # subpastas certas nunca seriam baixadas.
            sharepoint.baixar_todas_pastas(str(TMP_BASE), timeout=40)
            falhas = list(sharepoint.ULTIMAS_FALHAS)
            try:
                sem = semana_atual()
            except Exception:
                sem = None
            if sem:
                for s in (sem, sem - 1):
                    if s > 0:
                        sharepoint.baixar_rodada(s, str(TMP_BASE), timeout=30)
                        falhas += list(sharepoint.ULTIMAS_FALHAS)
            # Só marcamos como fresco se algo realmente chegou. Marcar após um
            # download vazio congelava a cópia velha por todo o TTL.
            if active_base() != TMP_BASE:
                print("⚠️ Nada foi baixado do SharePoint; mantendo a cópia empacotada.")
            elif falhas:
                # Parte das pastas não veio: os arquivos antigos continuam ali.
                # Marcamos um frescor curto para tentar de novo em ~1 min, em vez
                # de deixar o dado velho valendo o TTL inteiro.
                atraso = max(0, CACHE_TTL_SEGUNDOS - RETENTAR_APOS_FALHA)
                _MARKER.write_text(str(time.time() - atraso))
                print(f"⚠️ Download parcial ({', '.join(sorted(set(falhas)))}); "
                      f"nova tentativa em ~{RETENTAR_APOS_FALHA}s.")
            else:
                _MARKER.write_text(str(time.time()))
                print("✅ Dados do SharePoint atualizados em /tmp.")
        except Exception as e:
            print(f"⚠️ garantir_arquivos_frescos falhou: {e}")

# Metadados opcionais (nome amigável/tipo) por arquivo. NÃO define quais
# indicadores existem — os indicadores são descobertos automaticamente a partir
# dos arquivos .xlsx presentes nas pastas. Serve apenas como referência; o
# front-end mostra o nome do arquivo diretamente.
INDICADORES_MAP = {
    "VENDAS.xlsx": {"name": "Vendas", "type": "R$"},
    "PREMIER.xlsx": {"name": "Premier", "type": "R$"},
    "ELANCO.xlsx": {"name": "Antipulgas", "type": "R$"},
    "CAMAS ROUPAS COBERTORES.xlsx": {"name": "Suplementos", "type": "R$"},
    "LIMPEZA PERFUMARIA.xlsx": {"name": "Limpeza e Perfumaria", "type": "R$"},
}

# Pareamento explícito (opcional): força um arquivo da SEMANA ANTERIOR a casar
# com um arquivo da SEMANA ATUAL, caso o pareamento automático por similaridade
# não seja suficiente. Normalmente não é necessário — deixe vazio.
FILE_ALIASES = {}

# Quão parecidos dois nomes de arquivo precisam ser (0 a 1) para serem tratados
# como o MESMO indicador entre as semanas. Tolera prefixos ("MP "), espaços e
# pequenos erros de digitação.
SIMILARIDADE_MIN = 0.6


def _listar_xlsx(semana_path):
    """Lista os arquivos .xlsx de uma pasta (ignora temporários do Excel)."""
    if not semana_path.exists():
        return []
    return sorted(
        f for f in semana_path.glob("*.xlsx") if not f.name.startswith("~")
    )


def _chave(nome_arquivo):
    """Normaliza um nome de arquivo para comparação (maiúsculas, só alfanumérico)."""
    base = nome_arquivo.rsplit(".", 1)[0].upper()
    return re.sub(r"[^A-Z0-9]", "", base)


def _similaridade(nome_a, nome_b):
    return SequenceMatcher(None, _chave(nome_a), _chave(nome_b)).ratio()


def mapear_indicadores(semana=None):
    """Descobre os indicadores e pareia cada arquivo da SEMANA ATUAL com o
    arquivo correspondente da SEMANA ANTERIOR, mesmo que os nomes tenham
    pequenas diferenças. Retorna dict:
        { nome_arquivo_atual: {"anterior": Path|None, "atual": Path|None} }
    Quando um novo .xlsx é adicionado, ele entra automaticamente."""
    atual_files = _listar_xlsx(dir_atual(semana))
    anterior_files = _listar_xlsx(dir_anterior(semana))

    indicadores = {}
    for af in atual_files:
        indicadores[af.name] = {"anterior": None, "atual": af}

    restantes = list(anterior_files)

    # 1) Pareamento exato ou por alias explícito
    for canonico, slots in indicadores.items():
        for pf in list(restantes):
            if pf.name == canonico or FILE_ALIASES.get(pf.name) == canonico:
                slots["anterior"] = pf
                restantes.remove(pf)
                break

    # 2) Pareamento aproximado (tolera pequenas diferenças de nome)
    for canonico, slots in indicadores.items():
        if slots["anterior"] is not None:
            continue
        melhor, melhor_score = None, 0.0
        for pf in restantes:
            score = _similaridade(canonico, pf.name)
            if score > melhor_score:
                melhor, melhor_score = pf, score
        if melhor and melhor_score >= SIMILARIDADE_MIN:
            slots["anterior"] = melhor
            restantes.remove(melhor)

    # 3) Arquivos da semana anterior sem par viram indicadores próprios
    for pf in restantes:
        indicadores.setdefault(pf.name, {"anterior": pf, "atual": None})

    return indicadores


_TIPO_CACHE = {}


def detectar_tipo(file_path):
    """Tipo do indicador ('%' ou 'R$').

    Delega para calculo_rapido.detectar_tipo, que é a implementação usada no
    cálculo dos placares. Antes existiam duas cópias da mesma regra aqui e lá,
    e elas se separaram: quando o SHARE PIX passou a vir no layout longo (uma
    linha por loja por dia, com colunas de R$ ao lado da coluna de %), a cópia
    daqui olhava as colunas erradas e devolvia 'R$'. O detalhe do jogo então
    mostrava "R$ 0,17" em vez de "17,00%" e somava os dias em vez de usar o
    total da semana. Uma implementação só evita a divergência voltar.
    """
    if file_path is None:
        return "R$"
    chave = str(file_path)
    if chave in _TIPO_CACHE:
        return _TIPO_CACHE[chave]
    import calculo_rapido as cr
    tipo = cr.detectar_tipo(file_path)
    _TIPO_CACHE[chave] = tipo
    return tipo


def indicador_meta(arquivo, file_path=None, outro_path=None):
    """Nome/tipo do indicador. O tipo é DETECTADO do arquivo (formato da
    célula); INDICADORES_MAP serve só para o nome amigável."""
    import calculo_rapido as cr
    nome = INDICADORES_MAP.get(arquivo, {}).get("name") or cr.nome_limpo(arquivo)
    tipo = detectar_tipo(file_path) if file_path is not None else \
        INDICADORES_MAP.get(arquivo, {}).get("type", "R$")
    # 'nivel' = o gol vale pelo valor da própria semana (marcado no nome do
    # arquivo); 'evolucao' = regra padrão, evolução sobre a semana anterior.
    nomes = [arquivo] + [p.name for p in (file_path, outro_path) if p is not None]
    criterio = 'nivel' if any(cr.criterio_do_nome(n) == 'nivel' for n in nomes) else 'evolucao'
    return {"name": nome, "type": tipo, "criterio": criterio}


def ler_dias_loja(file_path, sigla):
    """Lê os valores dia a dia de uma loja. Retorna {dia: valor} ou None se a
    loja não estiver no arquivo.

    Usa o leitor de calculo_rapido, que tolera layouts diferentes (cabeçalho
    fora da 1ª linha, datas reais em vez de '10/08/2026 (Seg)', dias em ordem
    invertida e coluna 'Total') — caso do SHARE CLUBZ.
    """
    import calculo_rapido as cr
    try:
        dados = cr._carregar_arquivo(file_path)
    except Exception as e:
        print(f"⚠️ ler_dias_loja falhou ({file_path}): {e}")
        return None
    return dados.get(sigla)


def calcular_placar(team1, team2, semana):
    """Calcula o placar entre dois times somando a evolução dos indicadores.
    Retorna (score_team1, score_team2) onde cada ponto representa um indicador.

    Lógica:
    - Para cada indicador, compara a evolução total (semana atual vs anterior)
    - Time com maior evolução = 1 ponto naquele indicador
    - Soma dos 5 indicadores = placar final
    """
    placar_team1 = 0
    placar_team2 = 0

    try:
        # Obter dados de ambas as semanas para os dois times
        dados_team1 = {}
        dados_team2 = {}

        mapa = mapear_indicadores()
        for arquivo, slots in mapa.items():
            dados_team1[arquivo] = {}
            dados_team2[arquivo] = {}

            for semana_type in ("anterior", "atual"):
                file_path = slots.get(semana_type)
                if not file_path:
                    continue

                dias1 = ler_dias_loja(file_path, team1)
                dias2 = ler_dias_loja(file_path, team2)

                if dias1:
                    dados_team1[arquivo][semana_type] = dias1
                if dias2:
                    dados_team2[arquivo][semana_type] = dias2

        # Comparar evolução para cada indicador
        for arquivo in dados_team1.keys():
            if arquivo not in dados_team2:
                continue

            anterior1 = dados_team1[arquivo].get("anterior", {})
            atual1 = dados_team1[arquivo].get("atual", {})
            anterior2 = dados_team2[arquivo].get("anterior", {})
            atual2 = dados_team2[arquivo].get("atual", {})

            total_anterior1 = sum(anterior1.values()) if anterior1 else 0
            total_atual1 = sum(atual1.values()) if atual1 else 0
            evolucao1 = total_atual1 - total_anterior1

            total_anterior2 = sum(anterior2.values()) if anterior2 else 0
            total_atual2 = sum(atual2.values()) if atual2 else 0
            evolucao2 = total_atual2 - total_anterior2

            # Time com maior evolução ganha este indicador
            if evolucao1 > evolucao2:
                placar_team1 += 1
            elif evolucao2 > evolucao1:
                placar_team2 += 1
            # Se forem iguais, ninguém ganha ponto neste indicador

        return placar_team1, placar_team2

    except Exception as e:
        print(f"Erro ao calcular placar {team1} vs {team2}: {e}")
        return 0, 0

# ============================================================
# Utilitários para leitura de Excel
# ============================================================

def ler_arquivo_excel(file_path):
    """Lê arquivo Excel e retorna dict com dados por loja"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        dados = {}

        # Detectar quantas colunas de dados existem (a partir da coluna 3)
        header = list(ws.iter_rows(max_row=1, values_only=True))[0]
        num_cols = 2  # Começar da coluna C (índice 2)
        for col_idx in range(2, len(header)):
            if header[col_idx] and "202" in str(header[col_idx]):  # Procura por datas
                num_cols += 1
            else:
                break

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:  # Header
                continue

            loja = row[0]  # Primeira coluna
            if not loja:
                continue

            # Somar valores (colunas a partir de C até o último dia)
            valores_semana = []
            for col_idx in range(2, num_cols):
                val = row[col_idx]
                if val is not None:
                    try:
                        valores_semana.append(float(val))
                    except:
                        pass

            if valores_semana:
                total = sum(valores_semana)
                dados[loja] = total

        return dados

    except Exception as e:
        print(f"Erro ao ler {file_path}: {e}")
        return {}

def get_dados_indicadores(slot):
    """Lê todos os indicadores de uma semana (descobertos automaticamente).
    slot: "anterior" ou "atual"."""
    dados_semana = {}

    for arquivo, slots in mapear_indicadores().items():
        file_path = slots.get(slot)
        if not file_path:
            continue

        info = indicador_meta(arquivo, file_path)
        dados = ler_arquivo_excel(file_path)
        dados_semana[arquivo] = {
            "name": info["name"],
            "type": info["type"],
            "data": dados
        }

    return dados_semana

# ============================================================
# Endpoints da API
# ============================================================

@app.route('/api/health', methods=['GET'])
def health():
    """Verifica se a API está funcionando"""
    return jsonify({
        "status": "ok",
        "message": "API Campeonato Petz funcionando"
    })

# Rótulos da estrutura que não são regionais do campeonato: unidades de apoio
# (CD, hospital, hub, holding) e a linha placeholder marcada com "-". Elas
# entravam no filtro por regional e na busca por voz sem nunca ter jogo.
FORA_DO_CAMPEONATO = {'demaisunids', 'demaisunidades', 'holding', '-', ''}


def _fora_do_campeonato(*valores):
    for v in valores:
        chave = re.sub(r'[^a-z0-9]', '', str(v or '').strip().lower())
        if chave in FORA_DO_CAMPEONATO:
            return True
    return False


def estrutura_do_sharepoint():
    """Lê estrutura.xlsx (Regional | Distrito | Sigla Loja) da pasta raiz do
    SharePoint. Retorna {regional: {distrito: [lojas]}} ou None."""
    pasta = pasta_dados("Estrutura")
    arq = pasta / "estrutura.xlsx"
    if not arq.exists():
        cands = list(pasta.glob("*.xlsx")) if pasta.exists() else []
        if not cands:
            return None
        arq = cands[0]
    try:
        wb = openpyxl.load_workbook(arq, data_only=True, read_only=True)
        ws = wb.active
        linhas = list(ws.iter_rows(values_only=True))
        wb.close()
        est = {}
        for row in linhas[1:]:
            if not row or len(row) < 3:
                continue
            reg, dist, loja = row[0], row[1], row[2]
            if not reg or not dist or not loja:
                continue
            reg, dist, loja = str(reg).strip(), str(dist).strip(), str(loja).strip()
            if _fora_do_campeonato(reg, dist, loja):
                continue
            est.setdefault(reg, {}).setdefault(dist, []).append(loja)
        return est or None
    except Exception as e:
        print(f"⚠️ Falha ao ler estrutura do SharePoint: {e}")
        return None


def estrutura_ativa():
    """Estrutura vigente: SharePoint se disponível, senão estrutura.json."""
    est = estrutura_do_sharepoint()
    if est:
        return est
    try:
        import json as _json
        with open(Path(__file__).parent / 'estrutura.json') as f:
            return _json.load(f)
    except Exception:
        return {}


def nomes_lojas():
    """{sigla: nome da loja}, da coluna 'Nome Loja' do estrutura.xlsx.

    Serve ao comando de voz: ninguém fala "PDTP-SP", fala "Parada de Taipas".
    Sem essa coluna o reconhecimento cai na heurística sobre a sigla.
    """
    pasta = pasta_dados("Estrutura")
    arq = pasta / "estrutura.xlsx"
    if not arq.exists():
        cands = list(pasta.glob("*.xlsx")) if pasta.exists() else []
        if not cands:
            return {}
        arq = cands[0]
    try:
        wb = openpyxl.load_workbook(arq, data_only=True, read_only=True)
        ws = wb.active
        nomes = {}
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row or len(row) < 4:
                continue
            sigla, nome = row[2], row[3]
            if not sigla or not nome:
                continue
            sigla = str(sigla).strip()
            nome = ' '.join(str(nome).split())      # tira tabs e espaços duplos
            reg = row[0] if len(row) > 0 else None
            dist = row[1] if len(row) > 1 else None
            if _fora_do_campeonato(sigla, nome, reg, dist):
                continue
            nomes[sigla] = nome
        wb.close()
        return nomes
    except Exception as e:
        print(f"⚠️ Falha ao ler os nomes das lojas: {e}")
        return {}


@app.route('/api/lojas-nomes', methods=['GET'])
def get_lojas_nomes():
    """Nome por extenso de cada loja — usado pelo comando de voz."""
    try:
        garantir_arquivos_frescos()
    except Exception:
        pass
    n = nomes_lojas()
    return jsonify({"nomes": n, "total": len(n)})


@app.route('/api/estrutura', methods=['GET'])
def get_estrutura():
    """Estrutura Regional > Distrito > Lojas (do SharePoint, com fallback)."""
    try:
        garantir_arquivos_frescos()
    except Exception:
        pass
    est = estrutura_do_sharepoint()
    if est:
        return jsonify(est)
    try:
        import json as _json
        with open(Path(__file__).parent / 'estrutura.json') as f:
            return jsonify(_json.load(f))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _distritos_da_estrutura():
    """Lista de distritos da estrutura vigente (fonte da verdade do app)."""
    try:
        return [d for reg in estrutura_ativa().values() for d in reg]
    except Exception:
        return []


def _parear_com_estrutura(distritos_planilha):
    """Converte os nomes vindos do ranking para os nomes usados no app.
    Retorna (dict_pareado, nomes_da_planilha_sem_par, distritos_do_app_sem_historico).
    """
    import unicodedata

    def norm(s):
        s = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode().lower()
        return re.sub(r'[^a-z0-9]', '', s)

    def partes(nome):
        p = str(nome).split(' - ', 1)
        return norm(p[0]), (norm(p[1]) if len(p) > 1 else '')

    app_dists = _distritos_da_estrutura()
    if not app_dists:
        return distritos_planilha, [], []

    resultado, usados = {}, set()
    criterios = [
        lambda a, b: a == b,                                   # nome idêntico
        lambda a, b: partes(a)[0] == partes(b)[0],             # mesmo prefixo
        lambda a, b: partes(a)[1] and partes(a)[1] == partes(b)[1],  # mesma pessoa
    ]
    pendentes = list(distritos_planilha.keys())
    for cmp_ in criterios:
        for nome in list(pendentes):
            for alvo in app_dists:
                if alvo in usados:
                    continue
                if cmp_(nome, alvo):
                    resultado[alvo] = distritos_planilha[nome]
                    usados.add(alvo)
                    pendentes.remove(nome)
                    break
    sem_historico = [d for d in app_dists if d not in usados]
    return resultado, pendentes, sem_historico


def _parear_regionais(regionais_planilha):
    """Converte os nomes de regional do ranking para os nomes da estrutura.
    Casa por nome exato, prefixo (R1/R2/R3) ou nome da pessoa."""
    import unicodedata

    def norm(s):
        s = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode().lower()
        return re.sub(r'[^a-z0-9]', '', s)

    def partes(nome):
        p = str(nome).split(' - ', 1)
        return norm(p[0]), (norm(p[1]) if len(p) > 1 else '')

    alvos = list(estrutura_ativa().keys())
    if not alvos:
        return regionais_planilha, [], []
    resultado, usados = {}, set()
    criterios = [
        lambda a, b: a == b,
        lambda a, b: partes(a)[0] == partes(b)[0],
        lambda a, b: partes(a)[1] and partes(a)[1] == partes(b)[1],
    ]
    pendentes = list(regionais_planilha.keys())
    for cmp_ in criterios:
        for nome in list(pendentes):
            for alvo in alvos:
                if alvo in usados:
                    continue
                if cmp_(nome, alvo):
                    resultado[alvo] = regionais_planilha[nome]
                    usados.add(alvo)
                    pendentes.remove(nome)
                    break
    return resultado, pendentes, [a for a in alvos if a not in usados]


def historico_do_sharepoint(nome_pasta="Historico", chave="distrito"):
    """Lê o ranking acumulado das rodadas encerradas, do SharePoint.

    nome_pasta: 'Historico' (distritais) ou 'HistoricoRegional' (regionais).
    chave: 'distrito' ou 'regional' — define a coluna de identificação.

    Formato esperado (cabeçalho em qualquer uma das primeiras linhas):
        Rank | Distrital/Regional | Pontuação Média | Vit. Média
    O nº de rodadas vem do nome do arquivo ('rodada 6.xlsx'). Cada arquivo é o
    ranking ACUMULADO até aquela rodada. Retorna None se não houver planilha.
    """
    pasta = pasta_dados(nome_pasta)
    if not pasta.exists():
        return None
    # Só arquivos "rodada N" — a pasta pode conter outros itens por engano
    # (ex.: um 'Semana 8.xlsx' de confrontos salvo no lugar errado).
    def num_rodada(f):
        m = re.match(r"\s*rodada\s*(\d+)", f.stem, re.IGNORECASE)
        return int(m.group(1)) if m else -1

    arquivos = [f for f in pasta.glob("*.xlsx")
                if not f.name.startswith("~") and num_rodada(f) >= 0]
    if not arquivos:
        return None

    # A base é a rodada ANTERIOR à vigente (a atual ainda está em disputa e é
    # somada ao vivo). Se não existir, usa a maior rodada disponível.
    base_desejada = max(semana_atual() - 1, 1)
    candidatos = [f for f in arquivos if num_rodada(f) <= base_desejada]
    arq = max(candidatos or arquivos, key=num_rodada)

    def norm(s):
        # precisa remover ACENTOS, senão 'Pontuação Média' não é reconhecida
        import unicodedata
        s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode()
        return re.sub(r"[^a-z]", "", s.lower())

    try:
        wb = openpyxl.load_workbook(arq, data_only=True, read_only=True)
        ws = wb.active
        linhas = list(ws.iter_rows(values_only=True))
        wb.close()
        if not linhas:
            return None

        # Prefixos da coluna de identificação, conforme o tipo de ranking
        pref = (("regional",) if chave == "regional"
                else ("distrital", "distrito"))

        # Descobrir a linha de cabeçalho (a que tem a coluna de identificação)
        idx_head = None
        for i, row in enumerate(linhas[:10]):
            if any(norm(c).startswith(pref) for c in row):
                idx_head = i
                break
        if idx_head is None:
            return None
        head = linhas[idx_head]

        col_dist = col_pts = col_vit = col_rod = None
        for j, c in enumerate(head):
            n = norm(c)
            if col_dist is None and n.startswith(pref):
                col_dist = j
            elif col_pts is None and "pontuacaomedia" in n:
                col_pts = j
            elif col_vit is None and ("vitmedia" in n or "vitoriamedia" in n):
                col_vit = j
            elif col_rod is None and n.startswith("rodada"):
                col_rod = j
        if col_dist is None or col_pts is None:
            return None

        distritos = {}
        rodadas_col = None
        for row in linhas[idx_head + 1:]:
            if not row or col_dist >= len(row):
                continue
            nome = row[col_dist]
            if not nome or not str(nome).strip():
                continue
            try:
                pts = float(row[col_pts])
            except (TypeError, ValueError):
                continue
            vit = 0.0
            if col_vit is not None and col_vit < len(row):
                try:
                    vit = float(row[col_vit])
                except (TypeError, ValueError):
                    vit = 0.0
            if col_rod is not None and col_rod < len(row) and rodadas_col is None:
                try:
                    rodadas_col = int(float(row[col_rod]))
                except (TypeError, ValueError):
                    pass
            distritos[str(nome).strip()] = {
                "pontuacaoMedia": pts, "vitoriaMedia": vit
            }

        if not distritos:
            return None

        # Os nomes no ranking podem diferir dos da estrutura do app
        # (ex.: 'CO-1 - Ana B.' x 'CO-1 - Ana'). Pareia por: nome exato →
        # mesmo prefixo (CO-1, SP5...) → mesma pessoa (Jessica C.).
        if chave == "regional":
            distritos, nao_pareados, sem_historico = _parear_regionais(distritos)
        else:
            distritos, nao_pareados, sem_historico = _parear_com_estrutura(distritos)

        # Nº de rodadas: coluna > nome do arquivo > (semana vigente - 1)
        rodadas = rodadas_col
        if not rodadas:
            m = re.search(r"(\d+)", arq.stem)
            if m:
                rodadas = int(m.group(1))
        if not rodadas:
            rodadas = max(semana_atual() - 1, 1)

        return {
            "rodadasAnteriores": rodadas,
            "atualizadoEm": f"{arq.name} (SharePoint)",
            "distritos": distritos,
            "regionalDestaque": "R2 - Luiz",
            "origem": "sharepoint",
            "naoPareados": nao_pareados,      # nomes do ranking sem par no app
            "semHistorico": sem_historico,    # distritos do app sem histórico
        }
    except Exception as e:
        print(f"⚠️ Falha ao ler histórico do SharePoint: {e}")
        return None


@app.route('/api/historico', methods=['GET'])
def get_historico():
    """Histórico das rodadas encerradas. Prefere a planilha do SharePoint;
    se não houver, usa o historico.json empacotado."""
    try:
        garantir_arquivos_frescos()
    except Exception:
        pass
    # Ranking oficial das REGIONAIS (quando disponível, é a fonte preferida
    # para os totais por regional — evita derivar a partir dos distritos).
    hr = historico_do_sharepoint("HistoricoRegional", chave="regional")

    h = historico_do_sharepoint()
    if h:
        if hr:
            h["regionais"] = hr.get("distritos")
            h["rodadasRegionais"] = hr.get("rodadasAnteriores")
        return jsonify(h)
    try:
        import json as _json
        with open(Path(__file__).parent / 'historico.json') as f:
            d = _json.load(f)
        d["origem"] = "arquivo"
        if hr:
            d["regionais"] = hr.get("distritos")
            d["rodadasRegionais"] = hr.get("rodadasAnteriores")
        return jsonify(d)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================================
# CLASSIFICAÇÃO POR LOJA (grupos)
# ============================================================

def rodadas_classificacao():
    """Rodadas que têm arquivo de classificação por loja disponível."""
    base = pasta_dados("ClassificacaoLojas")
    if not base.exists():
        return []
    ns = set()
    for f in base.glob("*.xlsx"):
        if f.name.startswith("~"):
            continue
        m = re.search(r"(\d+)", f.stem)
        if m:
            ns.add(int(m.group(1)))
    return sorted(ns)


def classificacao_lojas(rodada=None):
    """Lê a classificação por loja/grupo exportada do Power BI.

    Pasta 'Classificação Lojas' no SharePoint, arquivos 'Rodada N.xlsx' com as
    colunas SERIE_GRUPO | Rank | Time | Pts | Jogos | VIT | EMP | DER | GM | GS | SG.
    Sem `rodada`, usa o arquivo da rodada mais alta; com `rodada`, usa aquele
    arquivo — é assim que dá para reabrir uma rodada passada, projetando a
    rodada N sobre a base da rodada N-1. Retorna (rodada, {grupo: [linhas]}).
    """
    base = pasta_dados("ClassificacaoLojas")
    if not base.exists():
        return None, {}

    melhor, melhor_n = None, -1
    for f in base.glob("*.xlsx"):
        if f.name.startswith("~"):
            continue
        m = re.search(r"(\d+)", f.stem)
        if not m:
            continue
        n = int(m.group(1))
        if rodada is not None:
            if n == int(rodada):
                melhor, melhor_n = f, n
                break
        elif n > melhor_n:
            melhor, melhor_n = f, n
    if not melhor:
        return None, {}

    wb = openpyxl.load_workbook(melhor, data_only=True, read_only=True)
    ws = wb.active
    grupos = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[2]:
            continue
        def _i(v):
            try:
                return int(v)
            except (TypeError, ValueError):
                return 0
        grupos.setdefault(str(row[0]).strip(), []).append({
            "rank": _i(row[1]), "time": str(row[2]).strip(),
            "pts": _i(row[3]), "jogos": _i(row[4]),
            "vit": _i(row[5]), "emp": _i(row[6]), "der": _i(row[7]),
            "gm": _i(row[8]), "gs": _i(row[9]), "sg": _i(row[10]),
        })
    wb.close()
    for linhas in grupos.values():
        linhas.sort(key=lambda r: r["rank"])
    return melhor_n, grupos


@app.route('/api/classificacao', methods=['GET'])
def get_classificacao():
    """Classificação por loja/grupo até a última rodada encerrada."""
    try:
        garantir_arquivos_frescos()
    except Exception:
        pass
    try:
        pedida = request.args.get('rodada', type=int)
        rodada, grupos = classificacao_lojas(pedida)
        return jsonify({
            "rodada": rodada,
            "grupos": grupos,
            "totalGrupos": len(grupos),
            "disponiveis": rodadas_classificacao(),
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ============================================================
# CALENDÁRIO COMPLETO (TODOS OS JOGOS.xlsx)
# ============================================================

def ler_todos_os_jogos():
    """Lê 'TODOS OS JOGOS.xlsx' da pasta Confrontos.

    Colunas: RODADA | ID_JOGO | MANDANTE | PLACAR | VISITANTE |
             GOLS_MANDANTE | GOLS_VISITANTE | STATUS
    Diferente dos 'Semana N.xlsx', este arquivo traz as 19 rodadas de uma vez
    e serve só para os insights (confronto direto, calendário futuro).
    """
    caminho = arquivo_confrontos("TODOS OS JOGOS.xlsx")
    if not caminho.exists():
        return []
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    ws = wb.active
    jogos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[2] or not row[4]:
            continue
        try:
            rodada = int(row[0])
        except (TypeError, ValueError):
            continue
        gm, gv = row[5], row[6]
        jogos.append({
            "rodada": rodada,
            "id": row[1],
            "mandante": str(row[2]).strip(),
            "visitante": str(row[4]).strip(),
            "golsMandante": gm if isinstance(gm, (int, float)) else None,
            "golsVisitante": gv if isinstance(gv, (int, float)) else None,
            "realizado": str(row[7] or "").strip().lower().startswith("realiz"),
        })
    wb.close()
    return jogos


@app.route('/api/jogos', methods=['GET'])
def get_jogos():
    """Calendário completo das 19 rodadas (para os insights)."""
    try:
        garantir_arquivos_frescos()
    except Exception:
        pass
    try:
        jogos = ler_todos_os_jogos()
        rodadas = sorted({j["rodada"] for j in jogos})
        return jsonify({
            "total": len(jogos),
            "rodadas": rodadas,
            "jogos": jogos,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ============================================================
# PRESENÇA: quem está online, de onde e desde quando
# ============================================================

MINUTOS_ONLINE = 10         # sem atividade por mais que isso = offline
INTERVALO_HEARTBEAT = 300   # só grava "visto agora" a cada N segundos
DIAS_HISTORICO_ACESSO = 30  # acessos mais antigos que isso são descartados


def _ip_do_cliente():
    """IP real do visitante. Atrás da Vercel o IP vem no X-Forwarded-For."""
    xff = request.headers.get('X-Forwarded-For', '')
    if xff:
        return xff.split(',')[0].strip()
    return request.remote_addr or '—'


def _local_do_cliente():
    """Cidade/região/país que a Vercel deduz do IP, sem chamar serviço externo."""
    from urllib.parse import unquote
    def h(nome):
        v = request.headers.get(nome, '')
        return unquote(v) if v else ''
    return h('x-vercel-ip-city'), h('x-vercel-ip-country-region'), h('x-vercel-ip-country')


def registrar_entrada(user):
    """Cria a sessão de acesso no login."""
    try:
        cidade, regiao, pais = _local_do_cliente()
        ac = Acesso(
            usuario_id=user.id, username=user.username, ip=_ip_do_cliente(),
            cidade=cidade, regiao=regiao, pais=pais,
            user_agent=(request.headers.get('User-Agent') or '')[:300],
        )
        db.session.add(ac)
        # Poda o histórico antigo na mesma transação: sem isso a tabela cresce
        # sem limite e o banco (plano gratuito) estoura a cota.
        corte = datetime.utcnow() - timedelta(days=DIAS_HISTORICO_ACESSO)
        Acesso.query.filter(Acesso.entrou_em < corte).delete(synchronize_session=False)
        db.session.commit()
        session['acesso_id'] = ac.id
    except Exception as e:
        db.session.rollback()
        print(f"⚠️ registrar_entrada falhou: {e}")


# Quando o banco está indisponível (ex.: cota do plano), paramos de tentar
# escrever presença por um tempo — cada tentativa custa uma conexão.
_presenca_pausada_ate = 0.0


@app.before_request
def _marcar_presenca():
    """Atualiza 'visto_em' da sessão atual, no máximo uma vez por minuto."""
    global _presenca_pausada_ate
    try:
        if not request.path.startswith('/api/'):
            return
        if not current_user.is_authenticated:
            return
        if time.time() < _presenca_pausada_ate:
            return
        agora = time.time()
        if agora - session.get('ultimo_ping', 0) < INTERVALO_HEARTBEAT:
            return
        session['ultimo_ping'] = agora
        aid = session.get('acesso_id')
        ac = db.session.get(Acesso, aid) if aid else None
        if ac is None:
            # Sessão anterior ao registro (ou perdida): abre uma nova.
            registrar_entrada(current_user)
            return
        ac.visto_em = datetime.utcnow()
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        _presenca_pausada_ate = time.time() + 600     # 10 min sem tentar
        print(f"⚠️ _marcar_presenca falhou: {e}")


@app.route('/api/acessos', methods=['GET'])
@login_required
def get_acessos():
    """Sessões online e histórico de entradas. Só para administradores."""
    if not getattr(current_user, 'é_admin', False):
        return jsonify({"error": "Apenas administradores"}), 403
    try:
        limite = datetime.utcnow() - timedelta(minutes=MINUTOS_ONLINE)

        def _iso_utc(dt):
            """ISO com fuso explícito. Sem isso o navegador interpretava a hora
            UTC como local e o horário aparecia 3h adiantado."""
            if dt is None:
                return None
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.isoformat()

        def fmt(a):
            partes = [p for p in (a.cidade, a.regiao, a.pais) if p]
            return {
                "username": a.username,
                "ip": a.ip,
                "local": " · ".join(partes) or "—",
                "dispositivo": _dispositivo(a.user_agent or ''),
                "entrouEm": _iso_utc(a.entrou_em),
                "vistoEm": _iso_utc(a.visto_em),
            }

        online = (Acesso.query.filter(Acesso.visto_em >= limite)
                  .order_by(Acesso.visto_em.desc()).all())
        historico = (Acesso.query.order_by(Acesso.entrou_em.desc()).limit(60).all())

        # Uma pessoa pode ter mais de uma aba/dispositivo: contamos usuários
        # distintos, não sessões.
        usuarios_online = sorted({a.username for a in online})

        return jsonify({
            "online": len(usuarios_online),
            "sessoesOnline": len(online),
            "usuarios": usuarios_online,
            "janelaMinutos": MINUTOS_ONLINE,
            "sessoes": [fmt(a) for a in online],
            "historico": [fmt(a) for a in historico],
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def _dispositivo(ua):
    """Resumo legível do User-Agent."""
    u = ua.lower()
    so = ('iPhone' if 'iphone' in u else 'iPad' if 'ipad' in u else
          'Android' if 'android' in u else 'Mac' if 'macintosh' in u else
          'Windows' if 'windows' in u else 'Linux' if 'linux' in u else '—')
    nav = ('Edge' if 'edg/' in u else 'Chrome' if 'chrome' in u and 'edg/' not in u else
           'Safari' if 'safari' in u else 'Firefox' if 'firefox' in u else '')
    return f"{so}{' · ' + nav if nav else ''}"


# ============================================================
# FAROL DOS INDICADORES: até que dia cada gol foi lançado
# ============================================================

@app.route('/api/farol/<int:semana>', methods=['GET'])
def get_farol(semana):
    """Último dia lançado em cada indicador da rodada.

    Verde = está no mesmo dia do indicador mais adiantado; vermelho = ficou
    para trás. Serve para ver de relance o que ainda falta subir.
    """
    try:
        garantir_arquivos_frescos()
        garantir_rodada(semana)
        import calculo_rapido as cr

        dias = cr.DIAS_ORDENADOS
        itens = []
        for arquivo, slots in mapear_indicadores(semana).items():
            fp = slots.get("atual")
            ultimo = -1
            if fp:
                try:
                    dados = cr._carregar_arquivo(fp)
                    for i, dia in enumerate(dias):
                        if any((v or {}).get(dia) for v in dados.values()):
                            ultimo = i
                except Exception as e:
                    print(f"⚠️ farol falhou em {arquivo}: {e}")
            itens.append({
                "indicador": cr.nome_limpo(arquivo),
                "ultimoDiaIdx": ultimo,
                "ultimoDia": dias[ultimo] if ultimo >= 0 else None,
            })

        referencia = max((i["ultimoDiaIdx"] for i in itens), default=-1)
        for i in itens:
            i["atualizado"] = i["ultimoDiaIdx"] == referencia and referencia >= 0

        itens.sort(key=lambda i: (i["atualizado"], i["indicador"]))
        return jsonify({
            "semana": semana,
            "referencia": dias[referencia] if referencia >= 0 else None,
            "indicadores": itens,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def historico_lojas():
    """Resultado de cada loja em cada rodada já encerrada.

    Duas fontes, nessa ordem de confiança:
      1) 'Classificação Lojas/Rodada N.xlsx' — a classificação oficial. A
         diferença entre duas rodadas consecutivas dá o resultado exato da
         rodada N (gols marcados, sofridos e pontos).
      2) 'Confrontos/TODOS OS JOGOS.xlsx' — o calendário com placares. Cobre as
         rodadas antigas, mas a rodada em andamento na hora da captura fica com
         placar parcial, então ela só entra onde a fonte 1 não alcança.

    Retorna {sigla: [{rodada, adv, gm, gs, res}]}, ordenado por rodada.
    """
    hist = {}

    # --- fonte 2: calendário completo (rodadas já realizadas) ---
    try:
        jogos_path = arquivo_confrontos("TODOS OS JOGOS.xlsx")
        if jogos_path.exists():
            wb = openpyxl.load_workbook(jogos_path, data_only=True, read_only=True)
            for row in wb.active.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 8 or row[0] is None:
                    continue
                rod, _id, man, _pl, vis, gm, gs, status = row[:8]
                if str(status or '').strip().lower() != 'realizado':
                    continue
                for time, adv, a, b in ((man, vis, gm, gs), (vis, man, gs, gm)):
                    if not time:
                        continue
                    hist.setdefault(str(time).strip(), {})[int(rod)] = {
                        "rodada": int(rod), "adv": str(adv).strip(),
                        "gm": int(a or 0), "gs": int(b or 0),
                        "res": 'V' if (a or 0) > (b or 0) else ('E' if (a or 0) == (b or 0) else 'D'),
                        "fonte": "calendario",
                    }
            wb.close()
    except Exception as e:
        print(f"⚠️ historico_lojas: calendário falhou ({e})")

    # --- fonte 1: diferença entre classificações oficiais consecutivas ---
    disponiveis = rodadas_classificacao()
    cache = {}
    for n in disponiveis:
        if n - 1 not in disponiveis:
            continue
        try:
            for r in (n - 1, n):
                if r not in cache:
                    _rod, grupos = classificacao_lojas(r)
                    cache[r] = {l["time"]: l for linhas in grupos.values() for l in linhas}
            ant, atu = cache[n - 1], cache[n]
            # Adversário daquela rodada, do arquivo de confrontos
            adv = {}
            conf_path = arquivo_confrontos(f"Semana {n}.xlsx")
            if conf_path.exists():
                import calculo_rapido as cr
                for c in cr.ler_confrontos(conf_path):
                    adv[c["team1"]] = c["team2"]
                    adv[c["team2"]] = c["team1"]
            for time, a in atu.items():
                b = ant.get(time)
                if not b or a["jogos"] - b["jogos"] != 1:
                    continue
                gm, gs = a["gm"] - b["gm"], a["gs"] - b["gs"]
                hist.setdefault(time, {})[n] = {
                    "rodada": n, "adv": adv.get(time, ''), "gm": gm, "gs": gs,
                    "res": 'V' if gm > gs else ('E' if gm == gs else 'D'),
                    "fonte": "oficial",
                }
        except Exception as e:
            print(f"⚠️ historico_lojas: rodada {n} falhou ({e})")

    return {t: [v[k] for k in sorted(v)] for t, v in hist.items()}


@app.route('/api/margens/<int:semana>', methods=['GET'])
@login_required
def get_margens(semana):
    """Por quanto cada gol da rodada foi ganho ou perdido.

    Alimenta o quadro "por pouco": quem perdeu um gol por cem reais e quem
    ganhou por menos ainda. O número é quanto faltava na SEMANA ATUAL para o
    perdedor alcançar a evolução do adversário — a mesma conta do "falta p/
    virar" que aparece no detalhe do jogo.
    """
    try:
        garantir_arquivos_frescos()
        import calculo_rapido as cr

        confrontos_path = arquivo_confrontos(f"Semana {semana}.xlsx")
        if not confrontos_path.exists():
            return jsonify({"error": f"Confrontos da semana {semana} não encontrados"}), 404
        confrontos = cr.ler_confrontos(confrontos_path)
        garantir_rodada(semana)
        memoria = cr.carregar_tudo(dir_anterior(semana), dir_atual(semana))
        if cr.semana_atual_vazia(memoria):
            return jsonify({"semana": semana, "semDados": True, "jogos": []})
        return jsonify({
            "semana": semana,
            "semDados": False,
            "jogos": cr.margens(confrontos, memoria),
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/api/historico-lojas', methods=['GET'])
@login_required
def get_historico_lojas():
    """Resultado rodada a rodada de cada loja — base do 'há quanto tempo não
    vence' e das sequências do resumo da mesa redonda."""
    try:
        garantir_arquivos_frescos()
        h = historico_lojas()
        ultima = max((j["rodada"] for v in h.values() for j in v), default=None)
        return jsonify({"lojas": h, "ultimaRodada": ultima, "total": len(h)})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/api/evolucao/<int:semana>', methods=['GET'])
@login_required
def get_evolucao(semana):
    """Placar de cada loja reconstruído ao final de cada dia da rodada.

    Alimenta o gráfico de desempenho da semana: com isso o front monta a curva
    de pontuação de cada distrito dia a dia e mostra em que dia (e em qual gol)
    a posição virou.
    """
    try:
        garantir_arquivos_frescos()
        import calculo_rapido as cr

        confrontos_path = arquivo_confrontos(f"Semana {semana}.xlsx")
        if not confrontos_path.exists():
            return jsonify({"error": f"Confrontos da semana {semana} não encontrados"}), 404
        confrontos = cr.ler_confrontos(confrontos_path)
        garantir_rodada(semana)
        memoria = cr.carregar_tudo(dir_anterior(semana), dir_atual(semana))
        ev = cr.evolucao_diaria(confrontos, memoria)
        ev["semana"] = semana
        ev["indicadores"] = sorted(memoria.keys())
        ev["criterios"] = {a: (m.get("criterio") or "evolucao") for a, m in memoria.items()}
        return jsonify(ev)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/api/semana', methods=['GET'])
def get_semana():
    """Semana vigente, detectada pelos arquivos de confronto disponíveis.
    Assim, ao subir 'Semana 6.xlsx' o site passa a usar a semana 6 sozinho."""
    try:
        garantir_arquivos_frescos()
    except Exception:
        pass
    return jsonify({
        "semana": semana_atual(),
        "disponiveis": semanas_disponiveis(),
        "origem": str(dir_confrontos()),
    })


@app.route('/api/diag', methods=['GET'])
def diagnostico():
    """Diagnóstico: mostra qual arquivo está sendo usado por indicador/semana,
    tamanho e quantos valores não-zero foram lidos. Ajuda a detectar arquivo
    desatualizado/corrompido em /tmp."""
    try:
        garantir_arquivos_frescos()
        pastas_tmp = {}
        for nome in ("SEMANA ANTERIOR", "SEMANA ATUAL", "Confrontos", "Historico", "Estrutura"):
            p = TMP_BASE / nome
            pastas_tmp[nome] = sorted(f.name for f in p.glob("*.xlsx")) if p.exists() else "AUSENTE"
        out = {
            "base_ativa": str(active_base()),
            "tmp_existe": (TMP_BASE / "SEMANA ATUAL").exists(),
            "idade_download_s": round(_idade_tmp()) if _idade_tmp() is not None else None,
            "pastas_tmp": pastas_tmp,
            "indicadores": {}
        }
        for arquivo, slots in mapear_indicadores().items():
            info = {}
            for semana in ("anterior", "atual"):
                fp = slots.get(semana)
                if not fp:
                    info[semana] = None
                    continue
                dados = {}
                try:
                    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
                    ws = wb.active
                    naozero = 0
                    total = 0
                    lojas = 0
                    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                        if not row or not row[0]:
                            continue
                        lojas += 1
                        for v in row[2:9]:
                            total += 1
                            if isinstance(v, (int, float)) and v != 0:
                                naozero += 1
                    wb.close()
                    dados = {"lojas": lojas, "celulas": total, "nao_zero": naozero}
                except Exception as e:
                    dados = {"erro": str(e)}
                info[semana] = {
                    "arquivo": fp.name,
                    "bytes": fp.stat().st_size if fp.exists() else None,
                    **dados
                }
            info["tipo"] = detectar_tipo(slots.get("atual") or slots.get("anterior"))
            out["indicadores"][arquivo] = info
        return jsonify(out)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/api/indicadores', methods=['GET'])
def get_indicadores():
    """Retorna lista de indicadores disponíveis"""
    return jsonify({
        "indicadores": [
            {"id": "VENDAS.xlsx", "name": "Vendas", "type": "R$"},
            {"id": "PREMIER.xlsx", "name": "Premier", "type": "R$"},
            {"id": "ELANCO.xlsx", "name": "Antipulgas", "type": "R$"},
            {"id": "CAMAS ROUPAS COBERTORES.xlsx", "name": "Suplementos", "type": "R$"},
            {"id": "MP LIMPEZA PERFUMARIA.xlsx", "name": "Share Marca Própria", "type": "%"},
            {"id": "LIMPEZA PERFUMARIA.xlsx", "name": "Úmidos Cães e Gatos", "type": "R$"},
        ]
    })

@app.route('/api/dados-semanas', methods=['GET'])
def get_dados_semanas():
    """Retorna todos os dados de semana anterior e atual"""
    try:
        dados_anterior = get_dados_indicadores("anterior")
        dados_atual = get_dados_indicadores("atual")

        return jsonify({
            "semana_anterior": dados_anterior,
            "semana_atual": dados_atual,
            "timestamp": "2026-07-22"
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/comparacao-lojas', methods=['POST'])
def comparacao_lojas():
    """
    Compara duas lojas e calcula evolução de indicadores
    Body: {"team1": "ABVT-SP", "team2": "ACST-SP"}
    """
    try:
        data = request.json
        team1 = data.get("team1")
        team2 = data.get("team2")

        if not team1 or not team2:
            return jsonify({"error": "Lojas inválidas"}), 400

        # Carregar dados
        dados_anterior = get_dados_indicadores("anterior")
        dados_atual = get_dados_indicadores("atual")

        resultado = {
            "team1": team1,
            "team2": team2,
            "indicadores": {},
            "gols": {"team1": 0, "team2": 0}
        }

        # Comparar cada indicador
        for arquivo, info in dados_anterior.items():
            data_anterior = info["data"]
            data_atual_info = dados_atual.get(arquivo, {})
            data_atual = data_atual_info.get("data", {})

            # Valores para team1
            val_ant_1 = data_anterior.get(team1, 0)
            val_atu_1 = data_atual.get(team1, 0)
            evolucao_1 = val_atu_1 - val_ant_1

            # Valores para team2
            val_ant_2 = data_anterior.get(team2, 0)
            val_atu_2 = data_atual.get(team2, 0)
            evolucao_2 = val_atu_2 - val_ant_2

            # Determinar vencedor do indicador
            vencedor = None
            if evolucao_1 > evolucao_2:
                vencedor = "team1"
                resultado["gols"]["team1"] += 1
            elif evolucao_2 > evolucao_1:
                vencedor = "team2"
                resultado["gols"]["team2"] += 1

            resultado["indicadores"][arquivo] = {
                "name": info["name"],
                "type": info["type"],
                "team1": {
                    "anterior": round(val_ant_1, 2),
                    "atual": round(val_atu_1, 2),
                    "evolucao": round(evolucao_1, 2)
                },
                "team2": {
                    "anterior": round(val_ant_2, 2),
                    "atual": round(val_atu_2, 2),
                    "evolucao": round(evolucao_2, 2)
                },
                "vencedor": vencedor
            }

        # Determinar resultado
        gols_1 = resultado["gols"]["team1"]
        gols_2 = resultado["gols"]["team2"]

        if gols_1 > gols_2:
            resultado["resultado"] = f"{team1} venceu {gols_1} x {gols_2}"
            resultado["pontos"] = {"team1": 3, "team2": 0}
        elif gols_2 > gols_1:
            resultado["resultado"] = f"{team2} venceu {gols_2} x {gols_1}"
            resultado["pontos"] = {"team1": 0, "team2": 3}
        else:
            resultado["resultado"] = f"Empate {gols_1} x {gols_2}"
            resultado["pontos"] = {"team1": 1, "team2": 1}

        return jsonify(resultado)

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/loja/<sigla>', methods=['GET'])
def get_loja(sigla):
    """Retorna dados de uma loja específica"""
    try:
        dados_anterior = get_dados_indicadores("anterior")
        dados_atual = get_dados_indicadores("atual")

        loja_data = {
            "sigla": sigla,
            "indicadores": {}
        }

        for arquivo, info in dados_anterior.items():
            loja_data["indicadores"][arquivo] = {
                "name": info["name"],
                "anterior": round(info["data"].get(sigla, 0), 2),
                "atual": round(dados_atual[arquivo]["data"].get(sigla, 0), 2)
            }

        return jsonify(loja_data)

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/confrontos/<int:semana>', methods=['GET'])
def get_confrontos(semana):
    """Retorna os confrontos de uma semana específica"""
    try:
        confrontos_path = arquivo_confrontos(f"Semana {semana}.xlsx")

        if not confrontos_path.exists():
            return jsonify({"error": f"Confrontos da semana {semana} não encontrados"}), 404

        wb = openpyxl.load_workbook(confrontos_path, data_only=True)
        ws = wb.active

        confrontos = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx <= 1:  # Skip header e linha vazia
                continue

            # Formato correto:
            # Col 0: DESC_JOGO_ELIMINATORIAS (sempre "Jogo")
            # Col 1: ID_JOGO (número)
            # Col 2: Time_Mandante
            # Col 3: Texto_Resultado_Partida_Card (ex: "1 x 5")
            # Col 4: Time_Visitante

            team1 = row[2]
            # score_str = row[3]  # Ignorar o score pré-determinado, será calculado
            team2 = row[4]

            if team1 and team2:
                confrontos.append({
                    "team1": team1,
                    "team2": team2
                })

        return jsonify({
            "semana": semana,
            "confrontos": confrontos,
            "total": len(confrontos)
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/loja-dias/<sigla>/<int:semana>', methods=['GET'])
def get_loja_dias(sigla, semana):
    """Retorna dados dia a dia de uma loja para uma semana"""
    try:
        garantir_arquivos_frescos()
        garantir_rodada(semana)
        dados_dias = {}

        # Indicadores descobertos automaticamente, já pareando o arquivo da
        # semana anterior com o da semana atual (tolerando pequenas diferenças
        # de nome).
        mapa = mapear_indicadores(semana)

        for arquivo, slots in mapa.items():
            # Tipo detectado do arquivo (prefere o da semana atual)
            info = indicador_meta(arquivo, slots.get("atual") or slots.get("anterior"),
                                  slots.get("anterior") if slots.get("atual") else None)

            for semana_type in ("anterior", "atual"):
                file_path = slots.get(semana_type)
                if not file_path:
                    continue

                dias = ler_dias_loja(file_path, sigla)
                if dias is None:
                    continue  # Loja não está neste arquivo

                dados_dias.setdefault(arquivo, {})[semana_type] = {
                    "name": info["name"],
                    "type": info["type"],
                    "criterio": info["criterio"],
                    "dias": dias
                }

        # Mapa de dias semana para índices (0=Seg, 6=Dom)
        dias_semana = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']
        hoje_idx = datetime.now().weekday()  # 0=Mon, 6=Sun
        # Converter para índice local (0=Seg brasileiro, 6=Dom)
        hoje_idx_br = (hoje_idx + 1) % 7  # Ajustar para semana começar em seg

        return jsonify({
            "sigla": sigla,
            "semana": semana,
            "dados": dados_dias,
            "hoje_dia": dias_semana[hoje_idx_br],
            "hoje_idx": hoje_idx_br
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/placar/<team1>/<team2>/<int:semana>', methods=['GET'])
def get_placar(team1, team2, semana):
    """Calcula o placar entre dois times baseado na evolução dos indicadores.
    Retorna score_team1 e score_team2, onde cada ponto = 1 indicador vencido."""
    try:
        score1, score2 = calcular_placar(team1, team2, semana)
        return jsonify({
            "team1": team1,
            "team2": team2,
            "score": f"{score1} x {score2}",
            "score1": score1,
            "score2": score2
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/lojas-disponiveis', methods=['GET'])
def get_lojas_disponiveis():
    """Retorna lista de todas as lojas disponíveis"""
    try:
        dados_anterior = get_dados_indicadores("anterior")

        # Pegar lojas do primeiro indicador
        primeiro_indicador = list(dados_anterior.values())[0]
        lojas = sorted(list(primeiro_indicador["data"].keys()))

        return jsonify({
            "total": len(lojas),
            "lojas": lojas
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ============================================================
# CALCULAR RESUMO DE TODOS OS JOGOS (PRÉ-CÁLCULO)
# ============================================================

def calcularPlacarBackend(team1, team2, semana, hojeIdx=None):
    """Calcula placar comparando evolução dos indicadores"""
    try:
        mapa = mapear_indicadores()
        score1 = 0
        score2 = 0
        diasOrdenados = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']

        for arquivo, slots in mapa.items():
            dados1_ind = {}
            dados2_ind = {}

            for semana_type in ("anterior", "atual"):
                file_path = slots.get(semana_type)
                if not file_path:
                    continue

                dias1 = ler_dias_loja(file_path, team1)
                dias2 = ler_dias_loja(file_path, team2)

                if dias1:
                    dados1_ind[semana_type] = dias1
                if dias2:
                    dados2_ind[semana_type] = dias2

            if not dados1_ind or not dados2_ind:
                continue

            dias1Anterior = dados1_ind.get('anterior', {})
            dias1Atual = dados1_ind.get('atual', {})
            dias2Anterior = dados2_ind.get('anterior', {})
            dias2Atual = dados2_ind.get('atual', {})

            diasAcontar = diasOrdenados[:hojeIdx+1] if hojeIdx is not None else diasOrdenados

            total1Anterior = sum(dias1Anterior.get(dia, 0) for dia in diasAcontar)
            total1Atual = sum(dias1Atual.get(dia, 0) for dia in diasAcontar)
            total2Anterior = sum(dias2Anterior.get(dia, 0) for dia in diasAcontar)
            total2Atual = sum(dias2Atual.get(dia, 0) for dia in diasAcontar)

            # Evolução PERCENTUAL (regra do campeonato) — casa com o frontend
            import calculo_rapido as _cr
            evolucao1 = _cr.evolucao_pct(total1Anterior, total1Atual)
            evolucao2 = _cr.evolucao_pct(total2Anterior, total2Atual)

            if evolucao1 > evolucao2:
                score1 += 1
            elif evolucao2 > evolucao1:
                score2 += 1
            # Evoluções iguais: desempate em cascata — maior valor atual,
            # depois maior valor da semana anterior
            elif total1Atual > total2Atual:
                score1 += 1
            elif total2Atual > total1Atual:
                score2 += 1
            elif total1Anterior > total2Anterior:
                score1 += 1
            elif total2Anterior > total1Anterior:
                score2 += 1

        return score1, score2
    except Exception as e:
        print(f"Erro ao calcular placar {team1} vs {team2}: {e}")
        return 0, 0

@app.route('/api/precalculate/<int:semana>', methods=['POST'])
def precalculate_games(semana):
    """Pré-calcula todos os jogos e salva em arquivo JSON para cache"""
    try:
        import json

        print(f"\n⏳ Iniciando pré-cálculo para semana {semana}...")

        # Carregar confrontos
        confrontos_path = arquivo_confrontos(f"Semana {semana}.xlsx")
        if not confrontos_path.exists():
            return jsonify({"error": f"Confrontos da semana {semana} não encontrados"}), 404

        wb = openpyxl.load_workbook(confrontos_path, data_only=True)
        ws = wb.active
        confrontos = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx <= 1:
                continue
            team1 = row[2]
            team2 = row[4]
            if team1 and team2:
                confrontos.append({"team1": team1, "team2": team2})

        games_summary = []
        print(f"⏳ Calculando {len(confrontos)} jogos...")

        # Pegar hoje_idx
        hoje_idx = 6
        try:
            hoje_idx_br = (datetime.now().weekday() + 1) % 7
            hoje_idx = hoje_idx_br
        except:
            pass

        for idx, conf in enumerate(confrontos):
            team1 = conf['team1']
            team2 = conf['team2']

            if (idx + 1) % 20 == 0:
                print(f"  {idx + 1}/{len(confrontos)} calculados...")

            score1_proj, score2_proj = calcularPlacarBackend(team1, team2, semana)
            score1_acum, score2_acum = calcularPlacarBackend(team1, team2, semana, hoje_idx)

            games_summary.append({
                "team1": team1,
                "team2": team2,
                "scoreProjected": f"{score1_proj} x {score2_proj}",
                "scoreAccumulated": f"{score1_acum} x {score2_acum}",
                "hojeIdx": hoje_idx
            })

        # Tentar salvar em arquivo (pode falhar em Vercel)
        try:
            cache_dir = BASE_PATH / "cache"
            cache_dir.mkdir(exist_ok=True)
            cache_file = str(cache_dir / f'games-summary-w{semana}.json')
            data = {
                "week": semana,
                "lastUpdated": datetime.now().isoformat(),
                "total": len(games_summary),
                "games": games_summary
            }
            with open(cache_file, 'w') as f:
                json.dump(data, f, indent=2)
            print(f"✅ {len(games_summary)} jogos calculados e salvos!")
        except Exception as cache_err:
            print(f"⚠️  Não foi possível salvar cache: {cache_err}")

        return jsonify({
            "message": f"Pré-cálculo concluído para semana {semana}",
            "week": semana,
            "total": len(games_summary),
            "games": games_summary
        }), 200

    except Exception as e:
        print(f"❌ Erro ao iniciar pré-cálculo: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

def _calcular_summary(semana):
    """Calcula o resumo de todos os jogos a partir da base ATIVA (/tmp se
    reprocessado, senão empacotada). Retorna o dict do resumo."""
    import calculo_rapido as cr
    confrontos_path = arquivo_confrontos(f"Semana {semana}.xlsx")
    if not confrontos_path.exists():
        # Distinguir "arquivo não existe" de "não conseguimos baixar": com o
        # SharePoint limitando as requisições (429) a pasta chega vazia e o
        # app cai para a cópia do repositório, que só tem semanas antigas.
        if _em_espera_throttle():
            raise FileNotFoundError(
                f"O SharePoint está limitando as requisições (erro 429) e os confrontos "
                f"da semana {semana} não puderam ser baixados. Tente de novo em alguns "
                f"minutos — os dados voltam sozinhos.")
        raise FileNotFoundError(f"Confrontos da semana {semana} não encontrados")
    confrontos = cr.ler_confrontos(confrontos_path)
    garantir_rodada(semana)          # dados da rodada pedida (subpasta)
    memoria = cr.carregar_tudo(dir_anterior(semana), dir_atual(semana))
    hoje_idx = (datetime.now().weekday() + 1) % 7
    jogos = cr.calcular_todos_jogos(confrontos, memoria, hoje_idx)

    # Alerta: indicador que subiu sem nenhum valor (planilha zerada). Sem base
    # de comparação o gol empata em todos os jogos e o placar não soma 6.
    avisos = []
    for arquivo, sem in memoria.items():
        nome = cr.nome_limpo(arquivo)
        nivel = sem.get("criterio") == "nivel"
        if nivel:
            avisos.append({
                "tipo": "criterio",
                "indicador": nome,
                "semana": f"rodada {semana}",
                "mensagem": f"{nome} está sendo disputado pelo valor da SEMANA ATUAL, "
                            f"não pela evolução: vence o gol quem tiver o maior número "
                            f"na semana. A base da semana anterior não é usada."
            })
        for rotulo, chave in (("semana anterior", "anterior"), ("semana atual", "atual")):
            lojas = sem.get(chave) or {}
            if not lojas:
                continue  # arquivo ausente nessa semana (não é o mesmo problema)
            if nivel and chave == "anterior":
                continue  # base intencionalmente sem uso — zerada não é erro
            tem_valor = any(v for dias in lojas.values()
                            for k, v in dias.items() if k != cr.CHAVE_TOTAL)
            if not tem_valor:
                avisos.append({
                    "tipo": "zerado",
                    "indicador": nome,
                    "semana": rotulo,
                    "mensagem": f"{nome} ({rotulo}) está sem dados — a planilha subiu zerada, "
                                f"então esse gol não está sendo disputado."
                })

    elim = sorted(cr.lojas_eliminadas())
    if elim:
        no_calendario = [e for e in elim
                         if any(e in (c["team1"].upper(), c["team2"].upper()) for c in confrontos)]
        if no_calendario:
            avisos.append({
                "tipo": "eliminada",
                "indicador": ", ".join(no_calendario),
                "semana": f"rodada {semana}",
                "mensagem": f"{', '.join(no_calendario)} está eliminada do campeonato: o resultado "
                            f"é administrativo, perde por 0 x 6 em todas as rodadas, "
                            f"independentemente da planilha de vendas."
            })

    rod_dados, _raiz = rodada_efetiva(semana)
    if rod_dados == semana and not _listar_xlsx(dir_atual(semana)):
        # Sem planilhas na SEMANA ATUAL da rodada. Duas situações diferentes:
        # a rodada vigente ainda não começou, ou uma rodada passada nunca teve
        # os arquivos arquivados na subpasta.
        tem_base = bool(_listar_xlsx(dir_anterior(semana)))
        passada = semana < semana_atual()
        if passada:
            msg = (f"A rodada {semana} não tem planilhas de venda arquivadas em "
                   f"'SEMANA ATUAL/rodada {semana}' no SharePoint, então não há "
                   f"resultados para exibir. Suba os arquivos dessa rodada para "
                   f"poder consultá-la.")
        elif tem_base:
            msg = (f"A rodada {semana} está começando: só há a base da semana "
                   f"anterior. Sem nenhum dia lançado na semana atual não há "
                   f"evolução para medir, então nenhum resultado está sendo "
                   f"contabilizado (sem vitória, empate ou derrota).")
        else:
            msg = (f"A rodada {semana} ainda não tem planilhas no SharePoint — "
                   f"nem a base da semana anterior, nem a semana atual.")
        avisos.append({
            "tipo": "rodada",
            "indicador": "Dados de venda",
            "semana": f"rodada {semana}",
            "mensagem": msg
        })
    if rod_dados != semana:
        avisos.append({
            "tipo": "rodada",
            "indicador": "Dados de venda",
            "semana": f"rodada {semana}",
            "mensagem": f"A rodada {semana} ainda não tem planilhas de venda — "
                        f"os placares estão sendo calculados com os dados da rodada {rod_dados}."
        })
    return {
        "week": semana,
        "rodadaDados": rod_dados,
        "semDadosAtual": cr.semana_atual_vazia(memoria),
        "eliminadas": sorted(cr.lojas_eliminadas()),
        "lastUpdated": datetime.now().isoformat(),
        "total": len(jogos),
        "games": jogos,
        "avisos": avisos,
    }


@app.route('/api/games-summary/<int:semana>', methods=['GET'])
def get_games_summary(semana):
    """Retorna o resumo de todos os jogos, sempre a partir dos dados mais
    frescos disponíveis (SharePoint via /tmp, com TTL)."""
    try:
        import json
        import os

        # 1) Garante que /tmp esteja fresco (rebaixa do SharePoint se preciso)
        garantir_arquivos_frescos()

        # 2) Calcular a partir da base ativa (rápido, ~0.4s)
        try:
            return jsonify(_calcular_summary(semana))
        except Exception as e:
            print(f"⚠️ Cálculo do resumo falhou ({e}). Tentando fallback empacotado.")

        # 3) Fallback: cache empacotado no repositório
        try:
            b_cache = str(BUNDLED_BASE / "cache" / f'games-summary-w{semana}.json')
            if os.path.exists(b_cache):
                with open(b_cache, 'r') as f:
                    return jsonify(json.load(f))
        except Exception:
            pass

        return jsonify({"error": "Não foi possível calcular o resumo"}), 500

    except Exception as e:
        print(f"❌ Erro ao retornar resumo: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ============================================================
# REPROCESSAR (baixa dados do SharePoint e recalcula)
# ============================================================

def _baixar_e_recalcular(semana):
    """FORÇA o download das pastas do SharePoint e recalcula. Usado pelo botão
    Reprocessar. Retorna dict com 'data' (resumo) e 'dias_atual'."""
    import calculo_rapido as cr

    # Força o refresh imediato (ignora o TTL) nesta instância
    garantir_arquivos_frescos(force=True)

    data = _calcular_summary(semana)

    # Dias disponíveis na semana atual (para informar o usuário)
    memoria = cr.carregar_tudo(dir_anterior(), dir_atual())
    dias_atual = []
    for arq, sem in memoria.items():
        for loja, dias in sem.get("atual", {}).items():
            dias_atual = list(dias.keys())
            break
        if dias_atual:
            break

    print(f"✅ Reprocessamento concluído: {data['total']} jogos.")
    return {"data": data, "dias_atual": dias_atual}


@app.route('/api/reprocessar/<int:semana>', methods=['POST'])
@login_required
def reprocessar(semana):
    """Baixa as pastas SEMANA ANTERIOR/ATUAL do SharePoint, recalcula todos os
    jogos e atualiza o cache. Retorna o resumo recalculado."""
    try:
        r = _baixar_e_recalcular(semana)
        return jsonify({
            "message": "Reprocessamento concluído com sucesso",
            "week": semana,
            "total": r["data"]["total"],
            "dias_semana_atual": r["dias_atual"],
            "lastUpdated": r["data"]["lastUpdated"],
        })
    except Exception as e:
        print(f"❌ Erro no reprocessamento: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Falha ao reprocessar: {e}"}), 500

# ============================================================
# AUTENTICAÇÃO
# ============================================================

@app.route('/api/login', methods=['POST'])
def login():
    """Login de usuário"""
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({"error": "Username e password são obrigatórios"}), 400

    # Com o banco fora, nem chegamos a consultar: tenta a emergência primeiro.
    emerg = autenticar_emergencia(username, password)
    if emerg is not None:
        login_user(emerg, remember=True)
        session.permanent = True
        return jsonify({"message": "Login em modo emergência", "user": emerg.to_dict()})

    try:
        user = Usuario.query.filter_by(username=username).first()
    except Exception as e:
        # Banco indisponível (ex.: plano suspenso). Tenta o acesso de
        # emergência, que não depende do Postgres.
        db.session.rollback()
        print(f"⚠️ login: banco indisponível ({e})")
        return jsonify({
            "error": "O banco de dados está indisponível no momento. "
                     "Use o acesso de emergência ou tente mais tarde."
        }), 503

    if user is None or not user.check_password(password):
        return jsonify({"error": "Username ou password inválidos"}), 401

    if not user.ativo:
        return jsonify({"error": "Usuário inativo"}), 403

    login_user(user, remember=True)
    session.permanent = True
    registrar_entrada(user)

    return jsonify({
        "message": "Login realizado com sucesso",
        "user": user.to_dict()
    })

@app.route('/api/logout', methods=['POST'])
@login_required
def logout():
    """Logout de usuário"""
    logout_user()
    return jsonify({"message": "Logout realizado com sucesso"})

@app.route('/api/me', methods=['GET'])
def get_current_user():
    """Retorna informações do usuário logado"""
    if current_user.is_authenticated:
        return jsonify({"user": current_user.to_dict()})
    return jsonify({"user": None})

@app.route('/api/usuarios', methods=['GET'])
@login_required
def list_usuarios():
    """Lista todos os usuários (apenas admin)"""
    if not current_user.é_admin:
        return jsonify({"error": "Acesso negado"}), 403

    usuarios = Usuario.query.all()
    return jsonify({
        "total": len(usuarios),
        "usuarios": [u.to_dict() for u in usuarios]
    })

@app.route('/api/usuarios', methods=['POST'])
@login_required
def create_usuario():
    """Cria novo usuário (apenas admin)"""
    if not current_user.é_admin:
        return jsonify({"error": "Acesso negado"}), 403

    data = request.get_json()
    username = data.get('username')
    email = data.get('email')
    nome_completo = data.get('nome_completo')
    password = data.get('password')

    if not username or not password:
        return jsonify({"error": "Username e password são obrigatórios"}), 400

    if Usuario.query.filter_by(username=username).first():
        return jsonify({"error": "Username já existe"}), 400

    if email and Usuario.query.filter_by(email=email).first():
        return jsonify({"error": "Email já existe"}), 400

    novo_usuario = Usuario(
        username=username,
        email=email,
        nome_completo=nome_completo,
        ativo=True
    )
    novo_usuario.set_password(password)
    db.session.add(novo_usuario)
    db.session.commit()
    invalidar_cache_usuarios()

    return jsonify({
        "message": "Usuário criado com sucesso",
        "user": novo_usuario.to_dict()
    }), 201

@app.route('/api/usuarios/<int:usuario_id>', methods=['PUT'])
@login_required
def update_usuario(usuario_id):
    """Atualiza usuário (admin) ou sua própria senha"""
    usuario = Usuario.query.get(usuario_id)

    if not usuario:
        return jsonify({"error": "Usuário não encontrado"}), 404

    # Usuário comum só pode atualizar sua própria senha
    if not current_user.é_admin and current_user.id != usuario_id:
        return jsonify({"error": "Acesso negado"}), 403

    data = request.get_json()

    # Atualizar senha
    if 'password' in data and data['password']:
        usuario.set_password(data['password'])

    # Apenas admin pode atualizar outros campos
    if current_user.é_admin:
        if 'email' in data:
            usuario.email = data['email']
        if 'nome_completo' in data:
            usuario.nome_completo = data['nome_completo']
        if 'ativo' in data:
            usuario.ativo = data['ativo']

    db.session.commit()
    invalidar_cache_usuarios()

    return jsonify({
        "message": "Usuário atualizado com sucesso",
        "user": usuario.to_dict()
    })

@app.route('/api/usuarios/<int:usuario_id>', methods=['DELETE'])
@login_required
def delete_usuario(usuario_id):
    """Deleta usuário (apenas admin)"""
    if not current_user.é_admin:
        return jsonify({"error": "Acesso negado"}), 403

    usuario = Usuario.query.get(usuario_id)

    if not usuario:
        return jsonify({"error": "Usuário não encontrado"}), 404

    if usuario.id == current_user.id:
        return jsonify({"error": "Não pode deletar a si mesmo"}), 400

    db.session.delete(usuario)
    db.session.commit()
    invalidar_cache_usuarios()

    return jsonify({"message": "Usuário deletado com sucesso"})

# ============================================================
# SERVIR ARQUIVOS ESTÁTICOS
# ============================================================

STATIC_DIR = str(Path(__file__).parent)

@app.route('/')
def index():
    """Serve a página principal do dashboard"""
    return send_from_directory(STATIC_DIR, 'dashboard-v3.html')

@app.route('/<path:filename>')
def serve_static(filename):
    """Serve arquivos estáticos (CSS, JS, etc)"""
    return send_from_directory(STATIC_DIR, filename)

# ============================================================
# Main
# ============================================================

if __name__ == '__main__':
    print("🚀 Iniciando Backend Campeonato Petz...")
    print(f"📁 Lendo dados de: {BASE_PATH}")
    print("🔗 API disponível em: http://localhost:5000")
    print("\nEndpoints:")
    print("  GET  /api/health")
    print("  GET  /api/indicadores")
    print("  GET  /api/dados-semanas")
    print("  GET  /api/lojas-disponiveis")
    print("  GET  /api/loja/<sigla>")
    print("  POST /api/comparacao-lojas")
    print("\n" + "="*60)

    app.run(debug=True, port=5000, host='0.0.0.0', threaded=True)
