"""
Cálculo otimizado dos jogos: carrega cada planilha UMA vez na memória e
calcula todos os confrontos a partir de dicionários — muito mais rápido que
abrir os arquivos repetidamente.

Este módulo é usado tanto pelo gerador de cache local quanto pelo endpoint de
reprocessamento no servidor.
"""
import os
import re
from pathlib import Path
from difflib import SequenceMatcher
import openpyxl

SIMILARIDADE_MIN = 0.6
FILE_ALIASES = {}
DIAS_ORDENADOS = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']

# Chave onde guardamos o valor da coluna 'Total' da planilha (não é um dia).
CHAVE_TOTAL = '__total__'


def _listar_xlsx(semana_path):
    if not semana_path.exists():
        return []
    return sorted(f for f in semana_path.glob("*.xlsx") if not f.name.startswith("~"))


def _chave(nome_arquivo):
    base = nome_arquivo.rsplit(".", 1)[0].upper()
    return re.sub(r"[^A-Z0-9]", "", base)


def _similaridade(nome_a, nome_b):
    return SequenceMatcher(None, _chave(nome_a), _chave(nome_b)).ratio()


def mapear_indicadores(semana_anterior, semana_atual):
    atual_files = _listar_xlsx(semana_atual)
    anterior_files = _listar_xlsx(semana_anterior)
    indicadores = {}
    for af in atual_files:
        indicadores[af.name] = {"anterior": None, "atual": af}
    restantes = list(anterior_files)
    for canonico, slots in indicadores.items():
        for pf in list(restantes):
            if pf.name == canonico or FILE_ALIASES.get(pf.name) == canonico:
                slots["anterior"] = pf
                restantes.remove(pf)
                break
    for canonico, slots in indicadores.items():
        if slots["anterior"] is not None:
            continue
        melhor, melhor_score = None, 0.0
        for pf in restantes:
            score = _similaridade(canonico, pf.name)
            if score > melhor_score:
                melhor, melhor_score = pf, score
        if melhor and melhor_score >= SIMILARIDADE_MIN:
            slots["anterior"] = melhor
            restantes.remove(melhor)
    for pf in restantes:
        indicadores.setdefault(pf.name, {"anterior": pf, "atual": None})
    return indicadores


def detectar_tipo(file_path):
    """Detecta se o indicador é percentual ('%') ou monetário ('R$').
    1) Formato 100% percentual -> '%'; 2) senão os VALORES decidem (fração
    -> '%', centenas/milhares -> 'R$'); 3) sem valores, usa o formato.
    Mantém a mesma regra de backend.detectar_tipo."""
    if file_path is None:
        return "R$"
    try:
        wb = openpyxl.load_workbook(file_path)  # precisa do formato
        ws = wb.active
        formatos = [str(c.number_format) for row in
                    ws.iter_rows(min_row=2, max_row=30, min_col=3, max_col=9)
                    for c in row if c.number_format]
        wb.close()
        com_pct = sum(1 for f in formatos if '%' in f)
        so_pct = bool(formatos) and com_pct == len(formatos)

        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active
        vals = [abs(v) for row in ws.iter_rows(min_row=2, min_col=3, max_col=9,
                                               values_only=True) if row
                for v in row if isinstance(v, (int, float)) and v != 0]
        wb.close()

        if so_pct:
            return "%"
        if vals:
            return "%" if all(v < 1 for v in vals) else "R$"
        return "%" if com_pct else "R$"
    except Exception as e:
        print(f"⚠️ detectar_tipo falhou ({file_path}): {e}")
    return "R$"


DIAS_PT = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']


def _dia_da_celula(valor):
    """Extrai o dia da semana ('Seg'...'Dom') de um cabeçalho de coluna.
    Aceita os dois formatos usados nas planilhas:
      - texto  '10/08/2026 (Seg)'   -> pega o que está entre parênteses
      - data   datetime(2026,8,10)  -> calcula o dia da semana
    Retorna None se a célula não representar um dia.
    """
    from datetime import datetime, date
    if valor is None:
        return None
    if isinstance(valor, (datetime, date)):
        return DIAS_PT[valor.weekday()]
    s = str(valor).strip()
    if '(' in s and ')' in s:
        dia = s[s.index('(') + 1:s.rindex(')')].strip()
        return dia if dia in DIAS_PT else None
    # texto de data ISO ou dd/mm/aaaa (sem o dia entre parênteses)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%Y %H:%M:%S"):
        try:
            return DIAS_PT[datetime.strptime(s, fmt).weekday()]
        except ValueError:
            continue
    return None


def _achar_cabecalho(linhas):
    """Encontra a linha de cabeçalho (a que tem mais colunas de dia) e as
    colunas de dias. Suporta planilhas com título/linhas em branco antes,
    como o SHARE CLUBZ (cabeçalho na 5ª linha). Ignora colunas 'Total'.
    A coluna 'Total' é devolvida à parte: em indicadores de share ela traz o
    percentual da semana calculado na origem (receita/receita), que NÃO é a
    média dos dias. Sua posição muda conforme os dias avançam.
    Retorna (indice_linha, [(col, dia)], col_total)."""
    melhor = (None, [], None)
    for i, row in enumerate(linhas[:15]):        # cabeçalho está nas 1ªs linhas
        if not row:
            continue
        cols = []
        col_total = None
        for j in range(1, len(row)):             # col 0 = sigla da loja
            if str(row[j] or '').strip().lower().startswith('total'):
                col_total = j                     # guardada, não vira dia
                continue
            dia = _dia_da_celula(row[j])
            if dia:
                cols.append((j, dia))
        # mantém só o primeiro de cada dia (evita duplicidade)
        vistos, únicos = set(), []
        for c, d in cols:
            if d not in vistos:
                vistos.add(d)
                únicos.append((c, d))
        if len(únicos) > len(melhor[1]):
            melhor = (i, únicos, col_total)
    return melhor


# Planilhas já lidas nesta instância, por (caminho, mtime, tamanho). Sem isso
# cada consulta de loja reabre os 12 arquivos da rodada — abrir a janela de um
# jogo relia 24 vezes o mesmo conteúdo.
_ARQ_CACHE = {}
_ARQ_CACHE_MAX = 48


def _guardar_cache(chave, dados):
    if chave is None:
        return
    if len(_ARQ_CACHE) >= _ARQ_CACHE_MAX:
        _ARQ_CACHE.pop(next(iter(_ARQ_CACHE)), None)
    _ARQ_CACHE[chave] = dados


def _chave_arquivo(file_path):
    try:
        st = os.stat(file_path)
        return (str(file_path), st.st_mtime_ns, st.st_size)
    except OSError:
        return None


def _carregar_arquivo(file_path):
    """Carrega TODAS as lojas de um arquivo de uma vez: {loja: {dia: valor}}.
    Tolera layouts diferentes (cabeçalho fora da 1ª linha, datas reais,
    dias em ordem invertida, coluna Total).

    O resultado fica em cache enquanto o arquivo não mudar (mtime/tamanho),
    então reprocessar ou baixar de novo do SharePoint invalida sozinho.
    """
    chave = _chave_arquivo(file_path)
    if chave is not None and chave in _ARQ_CACHE:
        return _ARQ_CACHE[chave]

    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active
    linhas = list(ws.iter_rows(values_only=True))
    wb.close()

    idx_head, col_dias, col_total = _achar_cabecalho(linhas)
    dados = {}
    if idx_head is None or not col_dias:
        _guardar_cache(chave, dados)
        return dados

    for row in linhas[idx_head + 1:]:
        if not row:
            continue
        sigla = row[0]
        if not sigla or not str(sigla).strip():
            continue
        dias = {}
        for col_idx, dia_nome in col_dias:
            valor = row[col_idx] if col_idx < len(row) else None
            try:
                if valor == "-" or valor is None:
                    dias[dia_nome] = 0
                else:
                    # 6 casas: preserva percentuais (0,21% = 0.0021).
                    # Arredondar em 2 casas zerava indicadores de share.
                    dias[dia_nome] = round(float(valor), 6)
            except (ValueError, TypeError):
                dias[dia_nome] = 0
        # Total da semana vindo da planilha (chave fora dos nomes de dia).
        if col_total is not None and col_total < len(row):
            try:
                v = row[col_total]
                if v not in (None, "-"):
                    dias[CHAVE_TOTAL] = round(float(v), 6)
            except (ValueError, TypeError):
                pass
        dados[str(sigla).strip()] = dias
    _guardar_cache(chave, dados)
    return dados


def carregar_tudo(semana_anterior, semana_atual):
    """Pré-carrega todos os indicadores em memória.
    Retorna {arquivo: {'anterior': {loja: {dia: val}}, 'atual': {...}}}."""
    mapa = mapear_indicadores(semana_anterior, semana_atual)
    memoria = {}
    for arquivo, slots in mapa.items():
        # Tipo detectado do arquivo (prefere a semana atual)
        tipo = detectar_tipo(slots.get("atual") or slots.get("anterior"))
        memoria[arquivo] = {"anterior": {}, "atual": {}, "tipo": tipo}
        for semana_type in ("anterior", "atual"):
            fp = slots.get(semana_type)
            if fp:
                memoria[arquivo][semana_type] = _carregar_arquivo(fp)
    return memoria


def agregar_pct(dias_obj, dias_a_contar):
    """Agregação de indicador percentual.

    Sempre a coluna 'Total' da planilha quando ela existe — em share é
    receita/receita, o número oficial do indicador. Vale inclusive com a semana
    incompleta: o total já reflete só os dias lançados. A média dos dias é
    apenas o plano B, para planilhas sem coluna de total.
    """
    o = dias_obj or {}
    total = o.get(CHAVE_TOTAL)
    if total:
        return total
    return _media_dias(o, dias_a_contar)


def _media_dias(dias_obj, dias_a_contar):
    """Média dos dias COM dado (ignora zeros = dia sem informação).
    Usado por indicadores percentuais, onde somar não faz sentido."""
    vals = [(dias_obj or {}).get(d, 0) for d in dias_a_contar]
    vals = [v for v in vals if v]
    return sum(vals) / len(vals) if vals else 0


def evolucao_pct(anterior, atual):
    """Evolução percentual usada no placar. Regras do campeonato:
      - sem base (semana anterior = 0)  -> 0% (não dá para medir evolução)
      - tinha valor e zerou nesta semana -> 0% (zero = sem dado, não queda de
        100%; evita punir a loja por lacuna na planilha)
      - caso normal -> (atual - anterior) / anterior * 100
    Precisa casar exatamente com evolucaoPct() no frontend.
    """
    if anterior == 0:
        return 0
    if atual == 0:
        return 0
    return (atual - anterior) / anterior * 100


def dias_com_dado(memoria):
    """Dias da semana ATUAL que já têm algum lançamento, em ordem.

    É o eixo do gráfico de evolução: só faz sentido reconstruir o placar até
    um dia que existe na planilha.
    """
    presentes = set()
    for semanas in memoria.values():
        for dias in (semanas.get("atual") or {}).values():
            for k, v in dias.items():
                if k != CHAVE_TOTAL and v:
                    presentes.add(k)
    return [d for d in DIAS_ORDENADOS if d in presentes]


def evolucao_diaria(confrontos, memoria):
    """Reconstrói o placar de cada confronto ao final de cada dia da rodada.

    Devolve, por loja, o resultado dela em cada gol dia a dia — a matéria-prima
    do gráfico de desempenho da semana e do resumo de oscilações.

    Formato: {'dias': [...], 'lojas': {sigla: {'adv': str,
              'gols': {arquivo: 'VDEV'}}}}, onde o caractere i é o resultado
    daquele gol considerando os lançamentos até o dia i.

    Ressalva dos indicadores percentuais: a planilha traz uma coluna 'Total'
    com o número oficial da semana, mas ela é uma foto do momento — não dá
    para recuperar o total de terça depois que quinta entrou. Nos dias
    intermediários usamos a média dos dias lançados; no último dia (o que
    aparece nas tabelas) usamos o Total, para o fim da curva bater exatamente
    com o placar exibido no site.
    """
    dias = dias_com_dado(memoria)
    lojas = {}
    if not dias:
        return {"dias": [], "lojas": lojas, "aproximaPct": False}

    ultimo = DIAS_ORDENADOS.index(dias[-1])
    aproxima = False
    for conf in confrontos:
        t1, t2 = conf["team1"], conf["team2"]
        e1 = lojas.setdefault(t1, {"adv": t2, "gols": {}})
        e2 = lojas.setdefault(t2, {"adv": t1, "gols": {}})
        for dia in dias:
            idx = DIAS_ORDENADOS.index(dia)
            # Só o último dia usa a coluna Total dos indicadores percentuais.
            # Placar PROJETADO do dia: é o que o site mostrava naquele dia —
            # a semana anterior inteira contra o que já tinha entrado.
            _s1, _s2, gols = _placar(memoria, t1, t2, idx,
                                     usar_total_pct=(idx == ultimo),
                                     truncar_anterior=False)
            for arquivo, vencedor in gols.items():
                e1["gols"].setdefault(arquivo, [])
                e2["gols"].setdefault(arquivo, [])
                e1["gols"][arquivo].append('V' if vencedor == 1 else ('D' if vencedor == 2 else 'E'))
                e2["gols"][arquivo].append('V' if vencedor == 2 else ('D' if vencedor == 1 else 'E'))
    for e in lojas.values():
        e["gols"] = {a: ''.join(v) for a, v in e["gols"].items()}
    for semanas in memoria.values():
        if semanas.get("tipo") == "%":
            aproxima = True
    return {"dias": dias, "lojas": lojas, "aproximaPct": aproxima and len(dias) > 1}


def _placar(memoria, team1, team2, hoje_idx=None, usar_total_pct=True,
            truncar_anterior=True):
    """Retorna (score1, score2, gols) onde gols é {arquivo: 1|2|0}:
    1 = team1 venceu o indicador, 2 = team2, 0 = empate."""
    score1 = score2 = 0
    gols = {}
    dias_a_contar = DIAS_ORDENADOS[:hoje_idx + 1] if hoje_idx is not None else DIAS_ORDENADOS
    # A semana anterior está sempre completa. Truncá-la junto (truncar_anterior)
    # dá a comparação dia a dia — Seg-Qui contra Seg-Qui — que é o placar
    # ACUMULADO. O placar PROJETADO, que alimenta as tabelas do site, compara a
    # semana anterior inteira com o que já entrou nesta semana.
    dias_anterior = dias_a_contar if truncar_anterior else DIAS_ORDENADOS
    for arquivo, semanas in memoria.items():
        ant = semanas["anterior"]
        atu = semanas["atual"]
        d1a, d1t = ant.get(team1), atu.get(team1)
        d2a, d2t = ant.get(team2), atu.get(team2)
        # Precisa dos dois times presentes no indicador
        if not (d1a or d1t) or not (d2a or d2t):
            continue
        # Indicador percentual (ex.: SHARE) agrega por MÉDIA dos dias com dado;
        # indicador monetário agrega por SOMA.
        ehPct = semanas.get("tipo") == "%"
        if ehPct:
            agg = (lambda o, ds: agregar_pct(o, ds)) if usar_total_pct else \
                  (lambda o, ds: _media_dias(o, ds))
        else:
            agg = (lambda o, ds: sum((o or {}).get(d, 0) for d in ds))
        t1_ant, t1_atu = agg(d1a, dias_anterior), agg(d1t, dias_a_contar)
        t2_ant, t2_atu = agg(d2a, dias_anterior), agg(d2t, dias_a_contar)
        # Evolução PERCENTUAL em relação à semana anterior (regra do campeonato).
        # Deve casar exatamente com calcularPlacarLocal() no frontend.
        ev1 = evolucao_pct(t1_ant, t1_atu)
        ev2 = evolucao_pct(t2_ant, t2_atu)
        if ev1 > ev2:
            score1 += 1
            gols[arquivo] = 1
        elif ev2 > ev1:
            score2 += 1
            gols[arquivo] = 2
        else:
            # Evoluções iguais (nenhum dos dois tem base, ou os dois zeraram).
            # Desempate em cascata: 1º maior valor na semana ATUAL;
            # 2º maior valor na semana ANTERIOR.
            if t1_atu > t2_atu:
                score1 += 1
                gols[arquivo] = 1
            elif t2_atu > t1_atu:
                score2 += 1
                gols[arquivo] = 2
            elif t1_ant > t2_ant:
                score1 += 1
                gols[arquivo] = 1
            elif t2_ant > t1_ant:
                score2 += 1
                gols[arquivo] = 2
            else:
                gols[arquivo] = 0   # empate real (sem dado em nenhuma semana)
    return score1, score2, gols


def semana_atual_vazia(memoria):
    """A semana ATUAL não tem nenhum valor em nenhum indicador?

    Acontece no primeiro dia da rodada, quando só existe a base de comparação.
    Sem nenhum dia lançado não há evolução para medir, então nenhum resultado
    deve ser atribuído — nem vitória, nem empate, nem derrota.
    """
    for semanas in memoria.values():
        for dias in (semanas.get("atual") or {}).values():
            if any(v for k, v in dias.items() if k != CHAVE_TOTAL):
                return False
    return True


def calcular_todos_jogos(confrontos, memoria, hoje_idx):
    """Calcula projetado + acumulado de todos os confrontos usando memória.
    Inclui 'golsProjetados' e 'golsAcumulados' (quem venceu cada indicador)."""
    jogos = []
    if semana_atual_vazia(memoria):
        # Rodada ainda não começou a receber vendas: jogos sem resultado.
        return [{
            "team1": c["team1"],
            "team2": c["team2"],
            "scoreProjected": "0 x 0",
            "scoreAccumulated": "0 x 0",
            "hojeIdx": hoje_idx,
            "golsProjetados": {},
            "golsAcumulados": {},
            "semDados": True,
        } for c in confrontos]
    for conf in confrontos:
        t1, t2 = conf["team1"], conf["team2"]
        s1p, s2p, gols_proj = _placar(memoria, t1, t2, None)
        s1a, s2a, gols_acum = _placar(memoria, t1, t2, hoje_idx)
        jogos.append({
            "team1": t1,
            "team2": t2,
            "scoreProjected": f"{s1p} x {s2p}",
            "scoreAccumulated": f"{s1a} x {s2a}",
            "hojeIdx": hoje_idx,
            "golsProjetados": gols_proj,
            "golsAcumulados": gols_acum,
        })
    return jogos


def ler_confrontos(confrontos_path):
    wb = openpyxl.load_workbook(confrontos_path, data_only=True, read_only=True)
    ws = wb.active
    confrontos = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx <= 1:
            continue
        team1 = row[2]
        team2 = row[4]
        if team1 and team2:
            confrontos.append({"team1": team1, "team2": team2})
    wb.close()
    return confrontos
